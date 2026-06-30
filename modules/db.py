import os
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime

# DATA_DIR pode ser sobrescrito por variável de ambiente (volume persistente no deploy).
_DATA_DIR = os.environ.get("DATA_DIR")
if _DATA_DIR:
    DB_PATH = Path(_DATA_DIR) / "financeiro.xlsx"
else:
    DB_PATH = Path(__file__).parent.parent / "data" / "financeiro.xlsx"
DB_PATH.parent.mkdir(parents=True, exist_ok=True)

# ── Backend de dados ──────────────────────────────────────────────────────────
# Se DATABASE_URL estiver definida (ex.: Supabase Postgres no deploy), usa Postgres.
# Caso contrário, usa o Excel local. Toda a aplicação passa por load_sheet/save_sheet,
# então a troca de backend é transparente para o restante do código.
_DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
USE_POSTGRES = bool(_DATABASE_URL)
_engine = None

def _normalize_pg_url(url: str) -> str:
    """Normaliza a URL do Postgres e codifica a senha (caracteres como @ na senha
    quebram o parse). Aceita senha crua ou já codificada."""
    from urllib.parse import quote, unquote
    url = url.strip()
    if url.startswith("postgres://"):
        url = url.replace("postgres://", "postgresql://", 1)
    if "://" not in url or "@" not in url:
        return url
    scheme, rest = url.split("://", 1)
    userinfo, host = rest.rsplit("@", 1)
    if ":" in userinfo:
        user, pwd = userinfo.split(":", 1)
        userinfo = f"{user}:{quote(unquote(pwd), safe='')}"
    return f"{scheme}://{userinfo}@{host}"


if USE_POSTGRES:
    from sqlalchemy import create_engine, inspect as _sa_inspect
    _engine = create_engine(_normalize_pg_url(_DATABASE_URL), pool_pre_ping=True)

# Colunas de cada tabela (usadas para criar tabelas vazias no Postgres)
_SCHEMA = {
    "lancamentos": ["id", "mes_ano", "cartao", "dono", "valor", "descricao",
                    "categoria", "valor_thais", "pessoa_thais", "tipo_parcela",
                    "parcela_atual", "total_parcelas", "id_grupo", "subtipo_cartao",
                    "data_lancamento", "conferido", "usuario"],
    "meses": ["mes_ano", "salario_kelvin", "salario_thais", "fechado", "usuario"],
    "fixos": ["id", "cartao", "descricao", "categoria", "valor_estimado",
              "pessoa_thais", "valor_thais", "ativo", "usuario"],
    "grupos_parcelamento": ["id", "descricao", "cartao", "subtipo_cartao", "categoria",
                            "valor_parcela", "total_parcelas", "mes_inicio",
                            "pessoa_thais", "valor_thais", "cancelado", "usuario"],
    "config": ["chave", "valor", "usuario"],
    "orcamentos": ["mes_ano", "categoria", "valor_planejado", "usuario"],
    "painel": ["mes_ano", "agua_boleto", "youtube_lembrete", "spotify_lembrete",
               "cdb_reserva", "previdencia", "usuario"],
    # Controles paralelos/manuais (Sprint 11, conta da Mãe): listas editáveis de
    # nome+valor por mês, usadas para 4 finalidades (campo `tipo`):
    #   "salario_componente" — itens que somados compõem o salário do mês (Aposent,
    #       Prefeit, Pensão...); o total alimenta automaticamente meses.salario_kelvin.
    #   "fixas"       — tabela livre de orçamento mensal (Cartão, Afepesp, Carro...),
    #                    exibida na página Lançamentos, sem efeito em outros cálculos.
    #   "restaurante" — tabela livre (nome usado como "dia"), mesma exibição.
    #   "aluguel"     — tabela livre; ao criar um mês novo, copia as linhas do mês
    #                    anterior automaticamente (ver upsert_mes).
    #   "cofrinho"    — extrato de poupança (descrição/valor/nota); também copiado
    #                    do mês anterior ao criar um mês novo (ela ajusta o que mudou).
    "controles_extra": ["id", "tipo", "mes_ano", "nome", "valor", "nota", "usuario"],
}

# Tabelas multi-usuário: cada conta (Kelvin, Mãe...) só vê suas próprias linhas.
# "usuario" não entra em queries por id (ids são globalmente únicos, ver _next_id).
_TABELAS_COM_USUARIO = set(_SCHEMA.keys())

SHEETS = ["lancamentos", "meses", "fixos", "config"]

# ── Multi-usuário ──────────────────────────────────────────────────────────────
# Conta atualmente logada (definida pelo app.py após autenticação). "kelvin" é o
# padrão/legado — todos os dados existentes antes da Sprint 11 pertencem a ele.
_usuario_atual = "kelvin"


def set_usuario_atual(usuario: str):
    global _usuario_atual
    _usuario_atual = usuario or "kelvin"


def get_usuario_atual() -> str:
    return _usuario_atual


def _next_id(sheet_name: str) -> int:
    """Próximo id, único entre TODAS as contas (evita colisão nas tabelas
    onde update/delete por id usa SQL direto sem filtrar por usuário)."""
    df = _load_sheet_raw(sheet_name)
    if df.empty or "id" not in df.columns or pd.isna(df["id"].max()):
        return 1
    return int(df["id"].max()) + 1

CATEGORIAS = ["Essencial", "Não essencial", "Estudos", "Lazer", "Viagem", "Reforma", "Negócios", "Metinha", "Livre"]
CARTOES = ["Santander", "Itaú", "C6", "Outros"]
SUBTIPOS_SANTANDER = []  # Santander sem subtipo — unificado
SUBTIPOS_ITAU = ["Visa", "Mastercard", "LATAM Pass"]
SUBTIPOS_ITAU_KELVIN = ["Visa", "Mastercard"]
TIPOS = ["única", "FIXO", "ULTIMA"]

CORES_CARTAO = {
    "Itaú":      {"bg": "#FF6B00", "text": "#FFFFFF"},
    "Santander": {"bg": "#EC0000", "text": "#FFFFFF"},
    "C6":        {"bg": "#242424", "text": "#F7C15F"},
    "Santander Físico": {"bg": "#A80000", "text": "#FFFFFF"},
    "Outros":    {"bg": "#16a34a", "text": "#FFFFFF"},
}


def db_exists():
    if USE_POSTGRES:
        return _sa_inspect(_engine).has_table("lancamentos")
    return DB_PATH.exists()


_CONFIG_PADRAO = [
    {"chave": "divisao_kelvin", "valor": "80"},
    {"chave": "divisao_thais", "valor": "20"},
    {"chave": "nome_kelvin", "valor": "Kelvin"},
    {"chave": "nome_thais", "valor": "Thais"},
]


def init_db():
    if USE_POSTGRES:
        # Cria tabelas vazias com as colunas corretas e popula config padrão.
        for nome, cols in _SCHEMA.items():
            if nome == "config":
                continue
            _write_sheet_raw(nome, pd.DataFrame(columns=cols))
        _write_sheet_raw("config", pd.DataFrame([{**c, "usuario": _usuario_atual} for c in _CONFIG_PADRAO]))
        return

    wb = Workbook()
    wb.remove(wb.active)
    for nome, cols in _SCHEMA.items():
        ws = wb.create_sheet(nome)
        ws.append(cols)
        if nome == "config":
            for c in _CONFIG_PADRAO:
                ws.append([c["chave"], c["valor"], _usuario_atual])
    wb.save(DB_PATH)


def _ensure_usuario_coluna_postgres(sheet_name: str):
    """Migração: garante a coluna `usuario` em tabelas Postgres antigas, preenchendo
    as linhas pré-existentes com 'kelvin' (todo dado anterior à conta da Mãe)."""
    from sqlalchemy import text, inspect as _sa_inspect2
    cols = [c["name"] for c in _sa_inspect2(_engine).get_columns(sheet_name)]
    if "usuario" in cols:
        return
    with _engine.begin() as conn:
        conn.execute(text(f'ALTER TABLE "{sheet_name}" ADD COLUMN usuario TEXT'))
        conn.execute(text(f'UPDATE "{sheet_name}" SET usuario = \'kelvin\' WHERE usuario IS NULL'))


def _load_sheet_raw(sheet_name: str) -> pd.DataFrame:
    """Lê a tabela inteira, sem filtrar por usuário (uso interno).

    IMPORTANTE: só retorna DataFrame vazio quando a tabela genuinamente não existe
    ainda (caso legítimo, ex.: 1ª inicialização). Qualquer outro erro de leitura
    (conexão caiu, timeout, etc.) é propagado (raise) — nunca tratado como "tabela
    vazia", pois save_sheet() usa este resultado para decidir quais linhas de
    OUTRAS contas preservar ao regravar a tabela inteira. Engolir o erro aqui já
    causou perda de dados real (apagou lançamentos de outra conta por engano)."""
    if USE_POSTGRES:
        if not _sa_inspect(_engine).has_table(sheet_name):
            return pd.DataFrame()
        if sheet_name in _TABELAS_COM_USUARIO:
            _ensure_usuario_coluna_postgres(sheet_name)
        return pd.read_sql_table(sheet_name, _engine)
    if not db_exists():
        return pd.DataFrame()
    try:
        df = pd.read_excel(DB_PATH, sheet_name=sheet_name, engine="openpyxl")
    except ValueError:
        return pd.DataFrame()  # aba não existe ainda no arquivo local
    if sheet_name in _TABELAS_COM_USUARIO and "usuario" not in df.columns:
        df["usuario"] = "kelvin"
    return df


def _write_sheet_raw(sheet_name: str, df: pd.DataFrame):
    """Escreve a tabela inteira (já com todas as contas), sem reaplicar filtro."""
    if USE_POSTGRES:
        df.to_sql(sheet_name, _engine, if_exists="replace", index=False)
        return
    wb = load_workbook(DB_PATH) if DB_PATH.exists() else Workbook()
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    elif "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        wb.remove(wb["Sheet"])
    ws = wb.create_sheet(sheet_name)
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))
    wb.save(DB_PATH)


def load_sheet(sheet_name: str) -> pd.DataFrame:
    df = _load_sheet_raw(sheet_name)
    if sheet_name in _TABELAS_COM_USUARIO and not df.empty and "usuario" in df.columns:
        df = df[df["usuario"] == _usuario_atual].copy()
    return df


def save_sheet(sheet_name: str, df: pd.DataFrame):
    """Recebe o conteúdo COMPLETO da conta atual para essa tabela (resultado de um
    load_sheet anterior, possivelmente modificado) e regrava preservando as linhas
    das demais contas, que não passam por aqui."""
    if sheet_name not in _TABELAS_COM_USUARIO:
        _write_sheet_raw(sheet_name, df)
        return
    df = df.copy()
    df["usuario"] = _usuario_atual
    outros = _load_sheet_raw(sheet_name)
    if not outros.empty and "usuario" in outros.columns:
        outros = outros[outros["usuario"] != _usuario_atual]
    else:
        outros = outros.iloc[0:0]
    final = pd.concat([outros, df], ignore_index=True) if not outros.empty else df
    _write_sheet_raw(sheet_name, final)


def get_config() -> dict:
    df = load_sheet("config")
    if df.empty:
        # Primeira vez dessa conta (ex.: Mãe logando antes de qualquer init_db
        # específico) — semeia a config padrão para não cair em todos os defaults
        # silenciosamente em todo lugar que usa cfg.get(...).
        save_sheet("config", pd.DataFrame(_CONFIG_PADRAO))
        df = load_sheet("config")
        if df.empty:
            return {}
    return dict(zip(df["chave"], df["valor"]))


def set_config(chave: str, valor):
    df = load_sheet("config")
    if chave in df["chave"].values:
        df.loc[df["chave"] == chave, "valor"] = str(valor)
    else:
        df = pd.concat([df, pd.DataFrame([{"chave": chave, "valor": str(valor)}])], ignore_index=True)
    save_sheet("config", df)


def get_meses() -> list[str]:
    df = load_sheet("meses")
    if df.empty:
        return []
    return sorted(df["mes_ano"].tolist())


def get_meses_fechados() -> set:
    """Conjunto de mes_ano marcados como fechados (já pagos)."""
    df = load_sheet("meses")
    if df.empty or "fechado" not in df.columns:
        return set()
    return set(df.loc[df["fechado"] == True, "mes_ano"].tolist())


def get_mes(mes_ano: str) -> dict:
    df = load_sheet("meses")
    row = df[df["mes_ano"] == mes_ano]
    if row.empty:
        return {"mes_ano": mes_ano, "salario_kelvin": 0.0, "salario_thais": 0.0, "fechado": False}
    return row.iloc[0].to_dict()


def upsert_mes(mes_ano: str, salario_kelvin: float, salario_thais: float, fechado: bool = None):
    """Atualiza salários do mês. `fechado` só é alterado se informado explicitamente —
    None preserva o status atual (evita reabrir o mês ao salvar só os salários)."""
    df = load_sheet("meses")
    mes_e_novo = mes_ano not in df["mes_ano"].values
    if not mes_e_novo:
        df.loc[df["mes_ano"] == mes_ano, "salario_kelvin"] = salario_kelvin
        df.loc[df["mes_ano"] == mes_ano, "salario_thais"] = salario_thais
        if fechado is not None:
            df.loc[df["mes_ano"] == mes_ano, "fechado"] = fechado
    else:
        novo = pd.DataFrame([{"mes_ano": mes_ano, "salario_kelvin": salario_kelvin,
                               "salario_thais": salario_thais, "fechado": bool(fechado)}])
        df = pd.concat([df, novo], ignore_index=True)
    save_sheet("meses", df)
    if mes_e_novo:
        for _tipo in ("aluguel", "cofrinho"):
            _copiar_controle_recorrente(_tipo, mes_ano)


_TIPOS_CONTROLE_RECORRENTE = ("aluguel", "cofrinho")


def _copiar_controle_recorrente(tipo: str, mes_ano_novo: str):
    """Ao criar um mês novo, copia uma tabela de controles_extra (nome/valor/nota)
    do mês anterior mais recente dessa conta — usado para 'aluguel' e 'cofrinho',
    controles manuais que naturalmente se repetem/acumulam mês a mês."""
    df = load_sheet("controles_extra")
    if df.empty:
        return
    anteriores = df[(df["tipo"] == tipo) & (df["mes_ano"] < mes_ano_novo)]
    if anteriores.empty:
        return
    ultimo_mes = anteriores["mes_ano"].max()
    if not df[(df["tipo"] == tipo) & (df["mes_ano"] == mes_ano_novo)].empty:
        return  # já tem dados desse tipo cadastrados nesse mês, não sobrescreve
    base = anteriores[anteriores["mes_ano"] == ultimo_mes]
    for _, r in base.iterrows():
        add_controle_extra(tipo, mes_ano_novo, str(r["nome"]),
                            float(r["valor"]) if pd.notna(r["valor"]) else 0.0,
                            str(r["nota"]) if pd.notna(r.get("nota")) else None)


def get_controles_extra(mes_ano: str, tipo: str) -> pd.DataFrame:
    df = load_sheet("controles_extra")
    if df.empty:
        return df
    return df[(df["mes_ano"] == mes_ano) & (df["tipo"] == tipo)].copy()


def add_controle_extra(tipo: str, mes_ano: str, nome: str, valor: float = 0.0, nota: str = None) -> int:
    df = load_sheet("controles_extra")
    novo_id = _next_id("controles_extra")
    novo = {"id": novo_id, "tipo": tipo, "mes_ano": mes_ano, "nome": nome,
            "valor": valor, "nota": nota}
    df = pd.concat([df, pd.DataFrame([novo])], ignore_index=True)
    save_sheet("controles_extra", df)
    if tipo == "salario_componente":
        _atualizar_salario_pelos_componentes(mes_ano)
    return novo_id


def update_controle_extra(item_id: int, **campos):
    df = load_sheet("controles_extra")
    if df.empty:
        return
    tipo = df.loc[df["id"] == item_id, "tipo"].iloc[0] if (df["id"] == item_id).any() else None
    for col, val in campos.items():
        if col in df.columns:
            df.loc[df["id"] == item_id, col] = val
    save_sheet("controles_extra", df)
    if tipo == "salario_componente":
        mes_ano = df.loc[df["id"] == item_id, "mes_ano"].iloc[0]
        _atualizar_salario_pelos_componentes(mes_ano)


def delete_controle_extra(item_id: int):
    df = load_sheet("controles_extra")
    if df.empty:
        return
    linha = df[df["id"] == item_id]
    tipo = linha["tipo"].iloc[0] if not linha.empty else None
    mes_ano = linha["mes_ano"].iloc[0] if not linha.empty else None
    df = df[df["id"] != item_id]
    save_sheet("controles_extra", df)
    if tipo == "salario_componente" and mes_ano:
        _atualizar_salario_pelos_componentes(mes_ano)


def _atualizar_salario_pelos_componentes(mes_ano: str):
    """Quando há itens em 'salario_componente' para o mês, o salário (salario_kelvin)
    passa a ser a soma deles automaticamente — usado pela conta da Mãe, cujo salário
    é composto por Aposentadoria + Prefeitura + Pensão + Conta etc."""
    soma = soma_controles_extra(mes_ano, "salario_componente")
    mes_atual = get_mes(mes_ano)
    upsert_mes(mes_ano, soma, float(mes_atual.get("salario_thais") or 0))


def soma_controles_extra(mes_ano: str, tipo: str) -> float:
    df = get_controles_extra(mes_ano, tipo)
    if df.empty:
        return 0.0
    return float(df["valor"].fillna(0).sum())


def set_mes_fechado(mes_ano: str, fechado: bool):
    """Marca/desmarca um mês como fechado (já pago), sem tocar nos salários."""
    df = load_sheet("meses")
    if mes_ano in df["mes_ano"].values:
        df.loc[df["mes_ano"] == mes_ano, "fechado"] = fechado
    else:
        novo = pd.DataFrame([{"mes_ano": mes_ano, "salario_kelvin": 0.0,
                               "salario_thais": 0.0, "fechado": fechado}])
        df = pd.concat([df, novo], ignore_index=True)
    save_sheet("meses", df)


def get_lancamentos(mes_ano: str) -> pd.DataFrame:
    df = load_sheet("lancamentos")
    if df.empty:
        return df
    return df[df["mes_ano"] == mes_ano].copy()


def add_lancamento(mes_ano, cartao, dono, valor, descricao, categoria,
                   valor_thais=None, pessoa_thais=None, tipo_parcela="única",
                   parcela_atual=None, total_parcelas=None, id_grupo=None,
                   subtipo_cartao=None, data_lancamento=None, conferido=False):
    from datetime import date as _date
    df = load_sheet("lancamentos")
    if "subtipo_cartao" not in df.columns:
        df["subtipo_cartao"] = None
    if "data_lancamento" not in df.columns:
        df["data_lancamento"] = None
    if "conferido" not in df.columns:
        df["conferido"] = False
    if data_lancamento is None:
        data_lancamento = _date.today().isoformat()
    novo_id = _next_id("lancamentos")
    novo = {
        "id": novo_id, "mes_ano": mes_ano, "cartao": cartao, "dono": dono,
        "valor": valor, "descricao": descricao, "categoria": categoria,
        "valor_thais": valor_thais, "pessoa_thais": pessoa_thais,
        "tipo_parcela": tipo_parcela, "parcela_atual": parcela_atual,
        "total_parcelas": total_parcelas, "id_grupo": id_grupo,
        "subtipo_cartao": subtipo_cartao, "data_lancamento": data_lancamento,
        "conferido": bool(conferido),
    }
    df = pd.concat([df, pd.DataFrame([novo])], ignore_index=True)
    save_sheet("lancamentos", df)
    return novo_id


def add_lancamentos_bulk(rows: list) -> int:
    """Insere vários lançamentos de uma vez (1 escrita só). Para criação de meses."""
    if not rows:
        return 0
    df = load_sheet("lancamentos")
    start_id = _next_id("lancamentos")
    prep = []
    for i, r in enumerate(rows):
        novo = {c: None for c in _SCHEMA["lancamentos"]}
        novo.update(r)
        novo["id"] = start_id + i
        novo["conferido"] = bool(novo.get("conferido"))
        prep.append(novo)
    novo_df = pd.DataFrame(prep)[_SCHEMA["lancamentos"]]
    df = pd.concat([df, novo_df], ignore_index=True) if not df.empty else novo_df
    save_sheet("lancamentos", df)
    return len(prep)


def get_painel(mes_ano: str) -> dict:
    """Campos editáveis do painel-resumo da Home (por mês)."""
    df = load_sheet("painel")
    base = {"mes_ano": mes_ano, "agua_boleto": 0.0, "youtube_lembrete": "",
            "spotify_lembrete": "", "cdb_reserva": 0.0, "previdencia": 0.0}
    if df.empty or "mes_ano" not in df.columns:
        return base
    row = df[df["mes_ano"] == mes_ano]
    if row.empty:
        return base
    d = row.iloc[0].to_dict()
    return {**base, **{k: v for k, v in d.items() if not pd.isna(v)}}


def set_painel(mes_ano: str, **campos):
    df = load_sheet("painel")
    cols = _SCHEMA["painel"]
    if df.empty or "mes_ano" not in df.columns:
        df = pd.DataFrame(columns=cols)
    if mes_ano in df["mes_ano"].values:
        for k, v in campos.items():
            if k in df.columns:
                df.loc[df["mes_ano"] == mes_ano, k] = v
    else:
        novo = {"mes_ano": mes_ano, **{c: None for c in cols if c != "mes_ano"}}
        novo.update(campos)
        df = pd.concat([df, pd.DataFrame([novo])], ignore_index=True)
    save_sheet("painel", df)


def calcular_painel(mes_ano: str) -> dict:
    """Reproduz as fórmulas do painel-resumo da planilha CONTAS.xlsx."""
    lanc = get_lancamentos(mes_ano)
    mes = get_mes(mes_ano)
    pin = get_painel(mes_ano)

    def _soma(df, col="valor"):
        return float(df[col].sum()) if not df.empty else 0.0

    if lanc.empty:
        cartao_k = cartao_t = pagamentos = mae = thais_cartao = 0.0
    else:
        cartao_k = _soma(lanc[lanc["dono"].astype(str).str.strip().str.lower() == "kelvin"])
        cartao_t = _soma(lanc[lanc["dono"].astype(str).str.strip().str.lower() == "thais"])
        vt = lanc["valor_thais"].fillna(0)
        pagamentos = -float(vt.sum())
        mae = -_soma(lanc[lanc["pessoa_thais"].astype(str).str.strip().str.lower() == "mãe"])
        mask_t = lanc["pessoa_thais"].astype(str).str.strip().str.lower() == "thais"
        thais_cartao = float(lanc.loc[mask_t, "valor_thais"].fillna(0).sum())

    agua = float(pin.get("agua_boleto") or 0)
    total_gastos = cartao_k + cartao_t + pagamentos + agua + mae
    sal_k = float(mes.get("salario_kelvin") or 0)
    sal_t = float(mes.get("salario_thais") or 0)
    total_thais = thais_cartao

    return {
        "salario_kelvin": sal_k, "salario_thais": sal_t,
        "cartao_kelvin": cartao_k, "cartao_thais": cartao_t,
        "pagamentos": pagamentos, "agua_boleto": agua, "mae": mae,
        "total_gastos": total_gastos,
        "diferenca_kelvin": sal_k - total_gastos,
        "thais_cartao": thais_cartao, "thais_total": total_thais,
        "diferenca_thais": sal_t - total_thais,
        "youtube_lembrete": pin.get("youtube_lembrete") or "",
        "spotify_lembrete": pin.get("spotify_lembrete") or "",
        "cdb_reserva": float(pin.get("cdb_reserva") or 0),
        "previdencia": float(pin.get("previdencia") or 0),
    }


def aplicar_fixos_ao_mes(mes_ano: str) -> int:
    """Sprint 2: cria lançamentos pendentes (valor=0) para cada fixo ativo no mês."""
    fixos = get_fixos(apenas_ativos=True)
    if fixos.empty:
        return 0
    lanc_existentes = get_lancamentos(mes_ano)
    descs_existentes = set(lanc_existentes["descricao"].str.strip().str.lower()) if not lanc_existentes.empty else set()
    adicionados = 0
    for _, fixo in fixos.iterrows():
        desc = str(fixo["descricao"]).strip()
        if desc.lower() in descs_existentes:
            continue
        add_lancamento(
            mes_ano=mes_ano, cartao=str(fixo["cartao"]), dono="Kelvin",
            valor=float(fixo.get("valor_estimado") or 0) or 0.0,
            descricao=desc, categoria=str(fixo["categoria"]),
            valor_thais=float(fixo["valor_thais"]) if fixo.get("valor_thais") and not pd.isna(fixo["valor_thais"]) else None,
            pessoa_thais=str(fixo["pessoa_thais"]) if fixo.get("pessoa_thais") and not pd.isna(fixo["pessoa_thais"]) else None,
            tipo_parcela="FIXO",
        )
        adicionados += 1
    return adicionados


def _proximo_mes(mes_ano: str, n: int) -> str:
    """Retorna mes_ano + n meses no formato YYYY-MM."""
    from datetime import date
    ano, mes = int(mes_ano[:4]), int(mes_ano[5:7])
    total = (mes - 1) + n
    return f"{ano + total // 12}-{(total % 12) + 1:02d}"


def criar_grupo_parcelamento(
    descricao, cartao, subtipo_cartao, categoria, valor_parcela, total_parcelas,
    mes_inicio, pessoa_thais=None, valor_thais=None,
) -> int:
    """Cria o grupo e todos os lançamentos mensais. Retorna o id do grupo."""
    df_g = load_sheet("grupos_parcelamento")
    if df_g.empty or "id" not in df_g.columns:
        df_g = pd.DataFrame(columns=["id", "descricao", "cartao", "subtipo_cartao",
                                      "categoria", "valor_parcela", "total_parcelas",
                                      "mes_inicio", "pessoa_thais", "valor_thais", "cancelado"])
    novo_gid = _next_id("grupos_parcelamento")
    novo_g = {
        "id": novo_gid, "descricao": descricao, "cartao": cartao,
        "subtipo_cartao": subtipo_cartao, "categoria": categoria,
        "valor_parcela": valor_parcela, "total_parcelas": total_parcelas,
        "mes_inicio": mes_inicio, "pessoa_thais": pessoa_thais,
        "valor_thais": valor_thais, "cancelado": False,
    }
    df_g = pd.concat([df_g, pd.DataFrame([novo_g])], ignore_index=True)
    save_sheet("grupos_parcelamento", df_g)

    for i in range(total_parcelas):
        mes = _proximo_mes(mes_inicio, i)
        # "Faltam" = parcelas restantes APÓS a atual (ex.: 4x → mês 1 mostra 3)
        faltam = total_parcelas - i - 1
        tipo = "ULTIMA" if faltam == 0 else "parcelado"
        # garante que o mês existe
        dados_mes = get_mes(mes)
        if not dados_mes.get("salario_kelvin"):
            upsert_mes(mes, 0.0, 0.0)
        add_lancamento(
            mes_ano=mes, cartao=cartao, dono="Kelvin",
            valor=valor_parcela, descricao=descricao, categoria=categoria,
            valor_thais=valor_thais, pessoa_thais=pessoa_thais,
            tipo_parcela=tipo, parcela_atual=i + 1,
            total_parcelas=faltam, id_grupo=novo_gid,
            subtipo_cartao=subtipo_cartao,
        )
    return novo_gid


def get_grupos_ativos() -> pd.DataFrame:
    """Retorna grupos com pelo menos uma parcela futura não cancelada."""
    df_g = load_sheet("grupos_parcelamento")
    if df_g.empty:
        return pd.DataFrame()
    df_l = load_sheet("lancamentos")
    if df_l.empty:
        return df_g
    # Para cada grupo, calcula parcelas restantes
    rows = []
    for _, g in df_g.iterrows():
        if g.get("cancelado"):
            continue
        lgs = df_l[df_l["id_grupo"] == g["id"]]
        total = int(g["total_parcelas"])
        pagas = len(lgs[lgs["tipo_parcela"].isin(["ULTIMA"])]) + len(
            lgs[(lgs["tipo_parcela"] == "parcelado") & (lgs["mes_ano"] < _mes_atual())]
        )
        restantes = total - pagas
        rows.append({**g.to_dict(), "pagas": pagas, "restantes": max(0, restantes)})
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def _mes_atual() -> str:
    from datetime import date
    d = date.today()
    return f"{d.year}-{d.month:02d}"


def cancelar_parcelas_restantes(grupo_id: int, mes_atual: str):
    """Remove lançamentos futuros de um grupo (antecipa quitação)."""
    df = load_sheet("lancamentos")
    df = df[~((df["id_grupo"] == grupo_id) & (df["mes_ano"] > mes_atual))]
    save_sheet("lancamentos", df)
    df_g = load_sheet("grupos_parcelamento")
    df_g.loc[df_g["id"] == grupo_id, "cancelado"] = True
    save_sheet("grupos_parcelamento", df_g)


def update_lancamento(lancamento_id: int, **campos):
    cols = [c for c in campos if c in _SCHEMA["lancamentos"]]
    if USE_POSTGRES and cols:
        # UPDATE direto (rápido) em vez de reescrever a tabela inteira
        from sqlalchemy import text
        sets = ", ".join(f'"{c}" = :{c}' for c in cols)
        params = {c: campos[c] for c in cols}
        params["_id"] = int(lancamento_id)
        params["_usuario"] = _usuario_atual
        with _engine.begin() as conn:
            conn.execute(text(
                f'UPDATE lancamentos SET {sets} WHERE id = :_id AND usuario = :_usuario'
            ), params)
        return
    df = load_sheet("lancamentos")
    for col, val in campos.items():
        if col in df.columns:
            df.loc[df["id"] == lancamento_id, col] = val
    save_sheet("lancamentos", df)


def corrigir_parcela_atual_grupo(id_grupo: int, mes_ano_de: str, parcela_inicial: int) -> int:
    """Resequencia parcela_atual a partir de `mes_ano_de`:
    o lançamento nesse mês fica com `parcela_inicial`, o seguinte +1, etc.
    Retorna quantos registros foram atualizados."""
    df = load_sheet("lancamentos")
    if df.empty:
        return 0
    mask = ((df["id_grupo"] == id_grupo)
            & (df["mes_ano"] >= mes_ano_de)
            & (df["tipo_parcela"].isin(["parcelado", "ULTIMA"])))
    ids_ordenados = df.loc[mask].sort_values("mes_ano")["id"].tolist()
    if not ids_ordenados:
        return 0
    if USE_POSTGRES:
        from sqlalchemy import text
        with _engine.begin() as conn:
            for i, rid in enumerate(ids_ordenados):
                conn.execute(text(
                    'UPDATE lancamentos SET parcela_atual = :pa WHERE id = :id AND usuario = :u'
                ), {"pa": parcela_inicial + i, "id": int(rid), "u": _usuario_atual})
    else:
        for i, rid in enumerate(ids_ordenados):
            df.loc[df["id"] == rid, "parcela_atual"] = parcela_inicial + i
        save_sheet("lancamentos", df)
    return len(ids_ordenados)


def propagar_parcela_grupo(id_grupo: int, mes_ano_de: str, **campos) -> int:
    """Aplica os mesmos campos (valor, valor_thais, pessoa_thais, categoria...) às demais
    parcelas (parcelado/ULTIMA) do grupo a partir de `mes_ano_de` (inclusive) — usado ao
    corrigir o valor de uma parcela já lançada, propagando para os meses seguintes.
    Também atualiza o registro do grupo (grupos_parcelamento) para refletir o novo valor."""
    df = load_sheet("lancamentos")
    if df.empty:
        return 0
    mask = ((df["id_grupo"] == id_grupo) & (df["mes_ano"] >= mes_ano_de)
            & (df["tipo_parcela"].isin(["parcelado", "ULTIMA"])))
    n = int(mask.sum())
    if n == 0:
        return 0
    if USE_POSTGRES:
        from sqlalchemy import text
        sets = ", ".join(f'"{c}" = :{c}' for c in campos)
        params = {**campos, "_grupo": int(id_grupo), "_mes": mes_ano_de, "_usuario": _usuario_atual}
        with _engine.begin() as conn:
            conn.execute(text(
                f'UPDATE lancamentos SET {sets} WHERE id_grupo = :_grupo AND mes_ano >= :_mes '
                f"AND tipo_parcela IN ('parcelado', 'ULTIMA') AND usuario = :_usuario"
            ), params)
    else:
        for col, val in campos.items():
            if col in df.columns:
                df.loc[mask, col] = val
        save_sheet("lancamentos", df)

    campos_grupo = {k: v for k, v in campos.items() if k in ("valor_parcela", "valor_thais", "pessoa_thais", "categoria")}
    if "valor" in campos:
        campos_grupo["valor_parcela"] = campos["valor"]
    if campos_grupo:
        df_g = load_sheet("grupos_parcelamento")
        if not df_g.empty and id_grupo in df_g["id"].values:
            for col, val in campos_grupo.items():
                if col in df_g.columns:
                    df_g.loc[df_g["id"] == id_grupo, col] = val
            save_sheet("grupos_parcelamento", df_g)
    return n


def delete_lancamento(lancamento_id: int):
    if USE_POSTGRES:
        from sqlalchemy import text
        with _engine.begin() as conn:
            conn.execute(text("DELETE FROM lancamentos WHERE id = :_id AND usuario = :_usuario"),
                         {"_id": int(lancamento_id), "_usuario": _usuario_atual})
        return
    df = load_sheet("lancamentos")
    df = df[df["id"] != lancamento_id]
    save_sheet("lancamentos", df)


def get_fixos(apenas_ativos: bool = True) -> pd.DataFrame:
    df = load_sheet("fixos")
    if df.empty:
        return df
    if apenas_ativos:
        return df[df["ativo"] == True].copy()
    return df.copy()


def update_fixo(fixo_id: int, **campos):
    df = load_sheet("fixos")
    for col, val in campos.items():
        if col in df.columns:
            df.loc[df["id"] == fixo_id, col] = val
    save_sheet("fixos", df)


def delete_fixo(fixo_id: int):
    df = load_sheet("fixos")
    df = df[df["id"] != fixo_id]
    save_sheet("fixos", df)


def resumo_mes(mes_ano: str) -> dict:
    lanc = get_lancamentos(mes_ano)
    mes = get_mes(mes_ano)
    if lanc.empty:
        total = 0.0
        por_categoria = {}
    else:
        total = float(lanc["valor"].sum())
        por_categoria = lanc.groupby("categoria")["valor"].sum().to_dict()
    sal_k = float(mes.get("salario_kelvin") or 0)
    sal_t = float(mes.get("salario_thais") or 0)
    return {
        "total_gasto": total,
        "salario_kelvin": sal_k,
        "salario_thais": sal_t,
        "saldo": (sal_k + sal_t) - total,
        "por_categoria": por_categoria,
    }


def fazer_backup(manter=7) -> str:
    """Copia o banco para data/backups/. Retorna o caminho do backup criado.
    No Postgres (Supabase), o backup é gerenciado pela própria plataforma — no-op."""
    if USE_POSTGRES:
        return ""
    import shutil
    from datetime import datetime as _dt
    if not db_exists():
        return ""
    backup_dir = DB_PATH.parent / "backups"
    backup_dir.mkdir(exist_ok=True)
    ts = _dt.now().strftime("%Y%m%d_%H%M%S")
    dest = backup_dir / f"financeiro_{ts}.xlsx"
    shutil.copy2(DB_PATH, dest)
    # Remove backups antigos, mantém os N mais recentes
    backups = sorted(backup_dir.glob("financeiro_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    for old in backups[manter:]:
        old.unlink()
    return str(dest)


def listar_backups() -> list:
    if USE_POSTGRES:
        return []
    backup_dir = DB_PATH.parent / "backups"
    if not backup_dir.exists():
        return []
    return sorted(backup_dir.glob("financeiro_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)


def get_orcamentos(mes_ano: str) -> dict:
    """Retorna {categoria: valor_planejado} para o mês."""
    df = load_sheet("orcamentos")
    if df.empty or "mes_ano" not in df.columns:
        return {}
    df = df[df["mes_ano"] == mes_ano]
    if df.empty:
        return {}
    return dict(zip(df["categoria"], df["valor_planejado"].astype(float)))


def set_orcamento(mes_ano: str, categoria: str, valor_planejado: float):
    df = load_sheet("orcamentos")
    if df.empty or "mes_ano" not in df.columns:
        df = pd.DataFrame(columns=["mes_ano", "categoria", "valor_planejado"])
    mask = (df["mes_ano"] == mes_ano) & (df["categoria"] == categoria)
    if mask.any():
        df.loc[mask, "valor_planejado"] = valor_planejado
    else:
        df = pd.concat([df, pd.DataFrame([{"mes_ano": mes_ano, "categoria": categoria,
                                            "valor_planejado": valor_planejado}])], ignore_index=True)
    save_sheet("orcamentos", df)


def delete_orcamento(mes_ano: str, categoria: str):
    df = load_sheet("orcamentos")
    if df.empty:
        return
    df = df[~((df["mes_ano"] == mes_ano) & (df["categoria"] == categoria))]
    save_sheet("orcamentos", df)


def calcular_divisao_mes(mes_ano: str, pct_k: float, pct_t: float) -> dict:
    """
    Retorna quanto cada um deve pagar no mês.

    Regra: se o lançamento tem valor_thais + pessoa_thais preenchidos,
    esse valor é responsabilidade de Pessoa (ex: Thais paga direto).
    O restante é dividido pct_k/pct_t entre Kelvin e a outra pessoa.
    """
    lanc = get_lancamentos(mes_ano)
    if lanc.empty:
        return {"total": 0.0, "kelvin": 0.0, "thais": 0.0, "thais_direto": 0.0,
                "saldo_kelvin": 0.0, "saldo_thais": 0.0}

    total = float(lanc["valor"].sum())

    # Valor direto de Pessoa (ex: Thais deve pagar de volta para Kelvin)
    mask_pessoa = lanc["pessoa_thais"].notna() & (lanc["valor_thais"].notna())
    thais_direto = float(lanc.loc[mask_pessoa, "valor_thais"].sum()) if mask_pessoa.any() else 0.0

    # Valor compartilhado = total menos itens que são 100% de Kelvin sem divisão
    # A divisão pct_k/pct_t é sobre o total gasto no cartão
    kelvin_paga = round(total * pct_k / 100, 2)
    thais_paga_proporcional = round(total * pct_t / 100, 2)

    return {
        "total": total,
        "kelvin": kelvin_paga,
        "thais_proporcional": thais_paga_proporcional,
        "thais_direto": thais_direto,
    }
