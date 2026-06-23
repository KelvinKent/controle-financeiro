from openpyxl import load_workbook
import pandas as pd
from modules.db import (
    DB_PATH, init_db, save_sheet, upsert_mes, load_sheet, get_fixos
)
from pathlib import Path

ORIGEM = Path(r"C:\Users\kelvin.araujo\Downloads\CONTAS.xlsx")

MESES_ALVO = [
    "Julho-26", "Agosto-26", "Setembro-26",
    "Outubro-26", "Novembro-26", "Dezembro-26",
]

MES_MAP = {
    "Julho-26": "2026-07", "Agosto-26": "2026-08", "Setembro-26": "2026-09",
    "Outubro-26": "2026-10", "Novembro-26": "2026-11", "Dezembro-26": "2026-12",
}

CARTOES_VALIDOS = {"Itaú", "Santander", "C6"}

NORMALIZA_CATEGORIA = {
    "viagen": "Viagem", "viagem": "Viagem",
    "essencial": "Essencial", "não essencial": "Não essencial",
    "estudos": "Estudos", "lazer": "Lazer", "reforma": "Reforma",
    "negócios": "Negócios", "metinha": "Metinha", "livre": "Livre",
}


def _normaliza_cat(val):
    if not val or not isinstance(val, str):
        return "Não essencial"
    return NORMALIZA_CATEGORIA.get(val.strip().lower(), val.strip())


def _parse_tipo(val):
    if val is None:
        return "única", None, None
    if isinstance(val, str):
        v = val.strip().upper()
        if v == "FIXO":
            return "FIXO", None, None
        if v == "ULTIMA":
            return "ULTIMA", None, None
        return "única", None, None
    if isinstance(val, (int, float)):
        total = int(val)
        return "parcelado", None, total
    return "única", None, None


def _extrair_fixos_template(wb):
    fixos = []
    visto = set()
    for nome_aba in MESES_ALVO:
        if nome_aba not in wb.sheetnames:
            continue
        ws = wb[nome_aba]
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            if not row or len(row) < 9:
                continue
            cartao = row[1]
            descricao = row[4]
            categoria = row[5]
            tipo_raw = row[8]
            if cartao not in CARTOES_VALIDOS:
                continue
            if isinstance(tipo_raw, str) and tipo_raw.strip().upper() == "FIXO":
                key = (cartao, descricao)
                if key not in visto and descricao:
                    visto.add(key)
                    valor_est = row[3] if isinstance(row[3], (int, float)) else 0
                    valor_t = row[6] if isinstance(row[6], (int, float)) else None
                    pessoa_t = row[7] if isinstance(row[7], str) else None
                    fixos.append({
                        "id": len(fixos) + 1,
                        "cartao": cartao,
                        "descricao": descricao,
                        "categoria": _normaliza_cat(categoria),
                        "valor_estimado": valor_est,
                        "pessoa_thais": pessoa_t,
                        "valor_thais": valor_t,
                        "ativo": True,
                    })
    return fixos


def migrar():
    if not ORIGEM.exists():
        raise FileNotFoundError(f"Planilha de origem não encontrada: {ORIGEM}")

    init_db()

    wb = load_workbook(ORIGEM, data_only=True)

    fixos = _extrair_fixos_template(wb)
    if fixos:
        df_fixos = pd.DataFrame(fixos)
        save_sheet("fixos", df_fixos)

    todos_lancamentos = []
    id_counter = 1

    for nome_aba in MESES_ALVO:
        if nome_aba not in wb.sheetnames:
            print(f"  Aba não encontrada: {nome_aba}")
            continue

        mes_ano = MES_MAP[nome_aba]
        ws = wb[nome_aba]

        sal_k = ws["K4"].value or 0
        sal_t = ws["N4"].value or 0
        upsert_mes(mes_ano, float(sal_k), float(sal_t))

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            if not row or len(row) < 9:
                continue
            cartao = row[1]
            dono = row[2]
            valor = row[3]
            descricao = row[4]
            categoria = row[5]
            valor_thais = row[6]
            pessoa_thais = row[7]
            tipo_raw = row[8]

            if cartao not in CARTOES_VALIDOS:
                continue
            if not isinstance(valor, (int, float)) or valor == 0:
                continue
            if not descricao:
                continue

            tipo, parc_atual, total_parc = _parse_tipo(tipo_raw)
            if isinstance(tipo_raw, (int, float)):
                total_parc = int(tipo_raw)

            cat = _normaliza_cat(categoria)
            vt = float(valor_thais) if isinstance(valor_thais, (int, float)) else None
            pt = str(pessoa_thais).strip() if isinstance(pessoa_thais, str) and pessoa_thais.strip() not in ("", "None") else None

            todos_lancamentos.append({
                "id": id_counter,
                "mes_ano": mes_ano,
                "cartao": str(cartao),
                "dono": str(dono) if dono else "Kelvin",
                "valor": float(valor),
                "descricao": str(descricao).strip(),
                "categoria": cat,
                "valor_thais": vt,
                "pessoa_thais": pt,
                "tipo_parcela": tipo,
                "parcela_atual": parc_atual,
                "total_parcelas": total_parc,
                "id_grupo": None,
            })
            id_counter += 1

        print(f"  {nome_aba} ({mes_ano}): {sum(1 for l in todos_lancamentos if l['mes_ano'] == mes_ano)} lançamentos | sal_k={sal_k} sal_t={sal_t}")

    if todos_lancamentos:
        df_lanc = pd.DataFrame(todos_lancamentos)
        save_sheet("lancamentos", df_lanc)

    print(f"\nMigração concluída: {len(todos_lancamentos)} lançamentos, {len(fixos)} fixos.")
    return len(todos_lancamentos)
