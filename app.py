import streamlit as st
import pandas as pd
import plotly.express as px
import sys
from pathlib import Path
from datetime import date

sys.path.insert(0, str(Path(__file__).parent))

from modules.db import (
    db_exists, get_meses, get_mes, upsert_mes, get_lancamentos,
    add_lancamento, update_lancamento, delete_lancamento, get_config, set_config,
    resumo_mes, CATEGORIAS, CARTOES, SUBTIPOS_SANTANDER, SUBTIPOS_ITAU, CORES_CARTAO,
    get_fixos, load_sheet, save_sheet, aplicar_fixos_ao_mes,
    criar_grupo_parcelamento, get_grupos_ativos, cancelar_parcelas_restantes,
    _proximo_mes, get_orcamentos, set_orcamento, delete_orcamento, calcular_divisao_mes,
    fazer_backup, listar_backups, update_fixo, delete_fixo,
    get_painel, set_painel, calcular_painel, add_lancamentos_bulk, set_mes_fechado,
    get_meses_fechados, propagar_parcela_grupo, set_usuario_atual, get_usuario_atual,
    get_controles_extra, add_controle_extra, delete_controle_extra,
)
from modules.fc_components import (
    inject_base_css, row_actions_css, bank_badge, hero_saldo,
    painel_resumo, painel_grid, lancamento_header, lancamento_row, card_cartao,
    CARTAO_COR, CARTAO_GRAD,
)

_MESES_PT = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
             "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
_MESES_PT_ABREV = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
                   "Jul", "Ago", "Set", "Out", "Nov", "Dez"]


def fmt_mes(ma, curto=False):
    """Formata 'AAAA-MM' como 'Mês/AAAA' (ou 'Mês/AA' abreviado). Qualquer ano."""
    try:
        a, m = str(ma).split("-")
        i = int(m) - 1
        if curto:
            return f"{_MESES_PT_ABREV[i]}/{a[2:]}"
        return f"{_MESES_PT[i]}/{a}"
    except Exception:
        return str(ma)


class _MesLabels:
    """Compatível com MES_LABELS.get(ma, default) — formata qualquer mês."""
    def get(self, ma, default=None):
        if not ma:
            return default if default is not None else ma
        return fmt_mes(ma)


MES_LABELS = _MesLabels()




def _fmt_desc(desc) -> str:
    """Normaliza a exibição da descrição: se estiver toda em maiúsculas
    (ex.: vindo de print de fatura), converte para formato Título.
    Não altera descrições já com capitalização própria."""
    s = str(desc) if desc is not None else ""
    letras = [c for c in s if c.isalpha()]
    if letras and all(c.isupper() for c in letras):
        return s.title()
    return s


def _editor_controle_extra(tipo: str, mes: str, titulo: str, label_nome: str = "Nome",
                            mostrar_nota: bool = False, ajuda: str = None):
    """Tabela editável genérica (nome/valor[/nota]) para os controles paralelos
    manuais da conta da Mãe (salário, fixas, restaurante, aluguel). 'Salvar' apaga
    e recria as linhas desse mês/tipo a partir do conteúdo editado."""
    import streamlit as _st
    df_ext = get_controles_extra(mes, tipo)
    cols = ["nome", "valor"] + (["nota"] if mostrar_nota else [])
    base = df_ext[cols].reset_index(drop=True) if not df_ext.empty else pd.DataFrame(columns=cols)
    col_config = {
        "nome": _st.column_config.TextColumn(label_nome),
        "valor": _st.column_config.NumberColumn("Valor (R$)", format="%.2f"),
    }
    if mostrar_nota:
        col_config["nota"] = _st.column_config.TextColumn("Nota")
    if ajuda:
        _st.caption(ajuda)
    edited = _st.data_editor(base, num_rows="dynamic", use_container_width=True,
                              column_config=col_config, key=f"editor_{tipo}_{mes}")
    if _st.button("💾 Salvar", key=f"salvar_{tipo}_{mes}", help=f"Salvar {titulo.lower()}"):
        for _, r in df_ext.iterrows():
            delete_controle_extra(int(r["id"]))
        for _, r in edited.iterrows():
            nome = str(r.get("nome") or "").strip()
            if not nome:
                continue
            valor = float(r.get("valor") or 0)
            nota = (str(r.get("nota")).strip() or None) if mostrar_nota and r.get("nota") else None
            add_controle_extra(tipo, mes, nome, valor, nota)
        if tipo == "salario_componente":
            _st.session_state.pop("sal_k", None)
        _st.success(f"{titulo} salvo(a)!")
        _st.rerun()
    total = float(edited["valor"].fillna(0).sum()) if not edited.empty and "valor" in edited.columns else 0.0
    _st.caption(f"Total: R$ {total:,.2f}")


def _mes_padrao(meses: list, fechados: set = frozenset()) -> str:
    """Primeiro mês (cronológico) ainda não fechado. Se todos estiverem fechados
    (ou não houver dados de fechamento), cai no mês mais recente."""
    if not meses:
        return "2026-07"
    for m in sorted(meses):
        if m not in fechados:
            return m
    return sorted(meses)[-1]


st.set_page_config(
    page_title="Controle Financeiro",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Tema centralizado via design system (fc_components) ──────────────────────
inject_base_css()
st.html(row_actions_css())
st.html("""
<style>
  .fc-card-label { font-size:13px; color:#aaa; margin-bottom:4px; }
  .fc-card-value { font-size:1.4rem; font-weight:700; }
  /* Impede quebra de texto em todos os botões */
  button[data-testid="baseButton-secondary"] p,
  button[data-testid="baseButton-primary"] p,
  button[data-testid="baseButton-destructive"] p { white-space: nowrap !important; }
</style>
""")


# Contas suportadas: id interno -> (nome de exibição, variável de ambiente/secrets da senha).
# Para adicionar uma nova conta (ex.: outro parente), basta acrescentar uma entrada aqui
# e cadastrar a respectiva senha (ver _ler_senha) — os dados de cada conta ficam isolados
# automaticamente (modules/db.py filtra tudo pela coluna "usuario").
_CONTAS = {
    "kelvin": {"nome": "Kelvin", "env": "APP_PASSWORD"},
    "mae": {"nome": "Mãe", "env": "APP_PASSWORD_MAE"},
}


def _ler_senha(env_key: str) -> str:
    import os as _os
    v = _os.environ.get(env_key, "")
    if not v:
        try:
            v = st.secrets[env_key]
        except Exception:
            v = ""
    return v


def _check_password() -> bool:
    """Gate de autenticação multi-conta. Cada conta em _CONTAS tem sua própria senha
    (variável de ambiente ou st.secrets). Conta sem senha configurada fica indisponível
    para login (exceto Kelvin: se nenhuma senha estiver configurada, libera direto, para
    manter o uso local sem senha como antes)."""
    senhas = {uid: _ler_senha(c["env"]) for uid, c in _CONTAS.items()}

    # Nenhuma senha configurada em lugar nenhum → acesso livre como Kelvin (uso local).
    if not any(senhas.values()):
        set_usuario_atual("kelvin")
        return True

    if st.session_state.get("_autenticado"):
        set_usuario_atual(st.session_state.get("_usuario", "kelvin"))
        return True

    contas_disponiveis = [uid for uid, s in senhas.items() if s]
    st.markdown("<div style='max-width:380px;margin:8vh auto 0'>", unsafe_allow_html=True)
    st.markdown("### 🔒 Controle Financeiro")
    st.caption("Acesso restrito. Selecione a conta e informe a senha para continuar.")
    if len(contas_disponiveis) > 1:
        usuario_sel = st.selectbox(
            "Conta", contas_disponiveis,
            format_func=lambda uid: _CONTAS[uid]["nome"],
        )
    else:
        usuario_sel = contas_disponiveis[0]
    senha = st.text_input("Senha", type="password", label_visibility="collapsed",
                          placeholder="Senha de acesso")
    if senha:
        if senha == senhas.get(usuario_sel):
            st.session_state["_autenticado"] = True
            st.session_state["_usuario"] = usuario_sel
            st.rerun()
        else:
            st.error("Senha incorreta.")
    st.markdown("</div>", unsafe_allow_html=True)
    return False


if not _check_password():
    st.stop()

if not db_exists():
    st.title("Configuração inicial")
    st.info("Banco de dados não encontrado. Executando migração da planilha original...")
    with st.spinner("Importando dados de Julho-26 a Dezembro-26..."):
        try:
            from modules.migration import migrar
            total = migrar()
            st.success(f"Migração concluída! {total} lançamentos importados.")
            st.rerun()
        except Exception as e:
            st.error(f"Erro na migração: {e}")
            st.stop()

meses_disponiveis = get_meses()

if not meses_disponiveis:
    # Conta nova (ex.: Mãe no primeiro acesso) sem nenhum mês cadastrado ainda —
    # cria o mês atual para a tela não ficar vazia/quebrada.
    _hoje = date.today()
    upsert_mes(f"{_hoje.year}-{_hoje.month:02d}", 0.0, 0.0)
    meses_disponiveis = get_meses()

# Backup automático diário — uma vez por sessão
if db_exists() and not st.session_state.get("_backup_feito"):
    from datetime import datetime as _dt
    backups = listar_backups()
    _hoje = _dt.now().strftime("%Y%m%d")
    _backup_hoje = any(_hoje in b.name for b in backups)
    if not _backup_hoje:
        fazer_backup()
    st.session_state["_backup_feito"] = True

if "mes_selecionado" not in st.session_state:
    st.session_state.mes_selecionado = _mes_padrao(meses_disponiveis, get_meses_fechados())

with st.sidebar:
    st.markdown("### 💰 Controle Financeiro")
    st.divider()
    labels = [MES_LABELS.get(m, m) for m in meses_disponiveis]
    idx = meses_disponiveis.index(st.session_state.mes_selecionado) if st.session_state.mes_selecionado in meses_disponiveis else 0
    escolha = st.selectbox("Mês", options=labels, index=idx)
    st.session_state.mes_selecionado = meses_disponiveis[labels.index(escolha)]
    st.divider()
    # Redirecionamento programático (ex.: após importar um "Fixo novo") precisa ser
    # resolvido ANTES do radio ser instanciado — não dá pra escrever na key depois.
    _redirect = st.session_state.pop("_nav_redirect", None)
    if _redirect:
        st.session_state["nav_pagina"] = _redirect
    pagina = st.radio("Navegação", ["Dashboard", "Histórico", "Lançamentos", "Parcelamentos", "Fixos", "Importar", "Configurações"], label_visibility="collapsed", key="nav_pagina")
    if st.session_state.get("_autenticado"):
        st.divider()
        st.caption(f"Conectado como **{_CONTAS.get(get_usuario_atual(), {}).get('nome', 'Kelvin')}**")
        if st.button("🚪 Sair", use_container_width=True):
            st.session_state["_autenticado"] = False
            st.session_state.pop("_usuario", None)
            st.session_state.pop("mes_selecionado", None)
            st.rerun()

mes = st.session_state.mes_selecionado
mes_label = MES_LABELS.get(mes, mes)
cfg = get_config()


# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
if pagina == "Dashboard":
    dados_mes = get_mes(mes)
    sal_k = float(dados_mes.get("salario_kelvin") or 15500)
    sal_t = float(dados_mes.get("salario_thais") or 6000)
    mes_fechado = bool(dados_mes.get("fechado")) if not pd.isna(dados_mes.get("fechado")) else False

    col_title, col_fechar = st.columns([3, 1])
    col_title.title(f"Dashboard — {mes_label}")
    with col_fechar:
        st.write("")
        if mes_fechado:
            st.success("🔒 Mês fechado")
            if st.button("🔓 Reabrir mês", use_container_width=True,
                         help="Volta a marcar este mês como em aberto."):
                set_mes_fechado(mes, False)
                st.rerun()
        else:
            if st.button("🔒 Fechar mês", use_container_width=True, type="primary",
                         help="Marca este mês como pago. Na próxima vez, o app já abre direto no mês seguinte."):
                set_mes_fechado(mes, True)
                meses_seguintes = [m for m in meses_disponiveis if m > mes]
                if meses_seguintes:
                    st.session_state.mes_selecionado = sorted(meses_seguintes)[0]
                st.success(f"{mes_label} fechado!")
                st.rerun()

    _usuario_mae = get_usuario_atual() == "mae"

    st.markdown("#### Salário do mês" if _usuario_mae else "#### Salários do mês")
    if _usuario_mae:
        col1, col3 = st.columns([3, 1])
        with col1:
            novo_sal_k = st.number_input(f"Salário {cfg.get('nome_kelvin','Rita')} (R$)",
                value=sal_k, min_value=0.0, step=100.0, format="%.2f", key="sal_k")
        novo_sal_t = 0.0
        with col3:
            st.write(""); st.write("")
            if st.button("💾", help="Salvar salário", use_container_width=True):
                upsert_mes(mes, novo_sal_k, novo_sal_t)
                st.success("Salário atualizado!")
                st.rerun()
    else:
        col1, col2, col3 = st.columns([2, 2, 1])
        with col1:
            novo_sal_k = st.number_input(f"Salário {cfg.get('nome_kelvin','Kelvin')} (R$)",
                value=sal_k, min_value=0.0, step=100.0, format="%.2f", key="sal_k")
        with col2:
            novo_sal_t = st.number_input(f"Salário {cfg.get('nome_thais','Thais')} (R$)",
                value=sal_t, min_value=0.0, step=100.0, format="%.2f", key="sal_t")
        with col3:
            st.write(""); st.write("")
            if st.button("💾", help="Salvar salários", use_container_width=True):
                upsert_mes(mes, novo_sal_k, novo_sal_t)
                st.success("Salários atualizados!")
                st.rerun()

    sal_k, sal_t = novo_sal_k, novo_sal_t

    if get_usuario_atual() == "mae":
        with st.expander("🧾 Componentes do salário (a soma vira o salário do mês acima)", expanded=False):
            _editor_controle_extra(
                "salario_componente", mes, "componentes do salário", label_nome="Componente",
                ajuda="Ex.: Aposentadoria, Prefeitura, Pensão, Conta... A soma destes itens "
                      "substitui automaticamente o campo \"Salário\" acima ao salvar.",
            )
        with st.expander("🐷 Cofrinho (poupança)", expanded=False):
            _editor_controle_extra(
                "cofrinho", mes, "cofrinho", label_nome="Descrição", mostrar_nota=True,
                ajuda="Extrato de poupança. Copiado automaticamente do mês anterior ao "
                      "criar um mês novo — só ajuste o que mudou (novo depósito, saque...).",
            )

    st.divider()

    # ── Painel-resumo (réplica da planilha CONTAS.xlsx) ───────────────────────
    # Específico do modelo Kelvin/Thais (Pix, reembolso da Mãe, YouTube/Spotify,
    # CDB/Previdência) — não se aplica à conta da Mãe, então o bloco inteiro
    # (grid "Resumo do mês" + o expander de edição) é pulado para ela.
    pin = calcular_painel(mes) if get_usuario_atual() != "mae" else None

    if pin is not None:
        nome_k = cfg.get('nome_kelvin', 'Kelvin')
        nome_t = cfg.get('nome_thais', 'Thais')
        st.markdown("#### Resumo do mês")
        st.html(painel_grid(
            painel_resumo(nome_k, [
                ("Salário", pin['salario_kelvin'], "kelvin"),
                ("Diferença", pin['diferenca_kelvin'], None),
            ]),
            painel_resumo(nome_t, [
                ("Salário", pin['salario_thais'], "thais"),
                ("Diferença", pin['diferenca_thais'], None),
            ]),
        ))
        st.html(painel_grid(
            painel_resumo("Gastos + Pix", [
                (f"Cartão {nome_k}", pin['cartao_kelvin'], None),
                (f"Cartão {nome_t}", pin['cartao_thais'], None),
                ("Pagamentos", pin['pagamentos'], None),
                ("Água - Boleto", pin['agua_boleto'], None),
                ("Mãe", pin['mae'], None),
                ("Total", pin['total_gastos'], None),
            ]),
            painel_resumo(f"Gastos {nome_t}", [
                ("Cartão", pin['thais_cartao'], None),
                ("Total", pin['thais_total'], None),
            ]),
        ))
        st.html(
            f'<div class="fc-box" style="font-family:inherit">'
            f'<div class="fc-hdr">Lembretes — quem me paga</div>'
            f'<div style="padding:8px 12px;font-size:13px"><b style="color:#ff5b5b">YouTube</b> — {pin["youtube_lembrete"] or "—"}</div>'
            f'<div style="padding:8px 12px;font-size:13px;border-top:1px solid rgba(255,255,255,0.06)"><b style="color:#1e9e5a">Spotify</b> — {pin["spotify_lembrete"] or "—"}</div>'
            f'</div>'
        )
        st.html(painel_resumo("Investimentos", [
            ("CDB (Reserva)", pin['cdb_reserva'], None),
            ("Previdência", pin['previdencia'], None),
        ], neg_red=False))

        with st.expander("⚙️ Editar campos do painel (Água, lembretes e investimentos)"):
            st.caption("Apenas estes campos são de digitação livre — os demais são calculados dos lançamentos.")
            pe1, pe2, pe3 = st.columns(3)
            ed_agua = pe1.number_input("Água - Boleto (R$)", min_value=0.0, step=1.0, format="%.2f",
                value=float(pin["agua_boleto"]), key="pin_agua")
            ed_cdb = pe2.number_input("CDB / Reserva (R$)", min_value=0.0, step=100.0, format="%.2f",
                value=float(pin["cdb_reserva"]), key="pin_cdb")
            ed_prev = pe3.number_input("Previdência (R$)", min_value=0.0, step=100.0, format="%.2f",
                value=float(pin["previdencia"]), key="pin_prev")
            ed_yt = st.text_input("Lembrete YouTube", value=pin["youtube_lembrete"], key="pin_yt")
            ed_sp = st.text_input("Lembrete Spotify", value=pin["spotify_lembrete"], key="pin_sp")
            if st.button("💾", key="btn_pin", help="Salvar painel"):
                set_painel(mes, agua_boleto=ed_agua, cdb_reserva=ed_cdb, previdencia=ed_prev,
                           youtube_lembrete=ed_yt, spotify_lembrete=ed_sp)
                st.success("Painel atualizado!")
                st.rerun()

        st.divider()

    res = resumo_mes(mes)
    total = res["total_gasto"]
    saldo = (sal_k + sal_t) - total

    def _card(label, value, nota=None, vermelho=False):
        cor = "#ff4b4b" if vermelho else "#21c354"
        nota_html = f'<div style="font-size:12px;color:{cor};margin-top:4px">{nota}</div>' if nota else ""
        return (f'<div class="fc-card-label">{label}</div>'
                f'<div class="fc-card-value">{value}</div>'
                f'{nota_html}')

    # ── Saldo combinado em destaque (hero) ─────────────────────────────────────
    st.html(hero_saldo(saldo))

    if _usuario_mae:
        c1, c3 = st.columns(2)
        c1.html(_card(f"Salário {cfg.get('nome_kelvin','Rita')}", f"R$ {sal_k:,.0f}"))
        c3.html(_card("Total gasto", f"R$ {total:,.0f}"))
    else:
        c1, c2, c3 = st.columns(3)
        c1.html(_card(f"Salário {cfg.get('nome_kelvin','Kelvin')}", f"R$ {sal_k:,.0f}"))
        c2.html(_card(f"Salário {cfg.get('nome_thais','Thais')}", f"R$ {sal_t:,.0f}"))
        c3.html(_card("Total gasto", f"R$ {total:,.0f}"))

    st.divider()

    lanc = get_lancamentos(mes)
    if lanc.empty:
        st.info("Nenhum lançamento neste mês ainda.")
    else:
        por_cat = lanc.groupby("categoria")["valor"].sum().sort_values(ascending=False)
        div_k = int(cfg.get("divisao_kelvin", 80))
        div_t = int(cfg.get("divisao_thais", 20))

        def _grafico_categorias():
            df_cat = por_cat.reset_index()
            df_cat.columns = ["Categoria", "Total (R$)"]
            df_cat["Total (R$)"] = df_cat["Total (R$)"].round(2)
            fig = px.bar(df_cat, x="Categoria", y="Total (R$)",
                text=df_cat["Total (R$)"].apply(lambda v: f"R$ {v:,.0f}"),
                color="Categoria", color_discrete_sequence=px.colors.qualitative.Set2)
            _ymax = float(df_cat["Total (R$)"].max() or 0)
            fig.update_layout(showlegend=False, margin=dict(t=90, b=10), height=340,
                              yaxis_range=[0, _ymax * 1.30], uniformtext_minsize=9,
                              uniformtext_mode="hide")
            fig.update_traces(textposition="outside", textfont_size=10, cliponaxis=False,
                              constraintext="none")
            st.plotly_chart(fig, use_container_width=True)

        if _usuario_mae:
            st.markdown("#### Gastos por categoria")
            _grafico_categorias()
        else:
            st.markdown(f"#### Gastos por categoria — divisão {div_k}/{div_t}")
            col_bar, col_div = st.columns([2, 1])
            with col_bar:
                _grafico_categorias()
            with col_div:
                st.markdown(f"**{cfg.get('nome_kelvin','Kelvin')}** ({div_k}%)")
                total_k = total * div_k / 100
                st.metric("Responsabilidade", f"R$ {total_k:,.0f}")
                saldo_k = sal_k - total_k
                st.metric("Saldo", f"R$ {saldo_k:,.0f}", delta_color="normal" if saldo_k >= 0 else "inverse")
                st.markdown(f"**{cfg.get('nome_thais','Thais')}** ({div_t}%)")
                total_t = total * div_t / 100
                st.metric("Responsabilidade", f"R$ {total_t:,.0f}")
                saldo_t = sal_t - total_t
                st.metric("Saldo", f"R$ {saldo_t:,.0f}", delta_color="normal" if saldo_t >= 0 else "inverse")

        st.divider()
        col_top, col_export = st.columns([3, 1])
        col_top.markdown("#### Top 10 maiores gastos variáveis")
        with col_export:
            st.write("")
            if st.button("📥 Excel", use_container_width=True, help="Baixar lançamentos do mês como .xlsx"):
                import io
                from openpyxl import Workbook as _WB
                exp = lanc.copy()
                exp = exp.drop(columns=["data_lancamento"], errors="ignore")
                wb_exp = _WB()
                ws_exp = wb_exp.active
                ws_exp.title = mes_label.replace("/", "-")
                ws_exp.append(list(exp.columns))
                for r in exp.itertuples(index=False):
                    ws_exp.append(list(r))
                buf = io.BytesIO()
                wb_exp.save(buf)
                buf.seek(0)
                st.download_button("⬇️ Baixar", data=buf,
                    file_name=f"lancamentos_{mes}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        variaveis = lanc[~lanc["tipo_parcela"].isin(["FIXO"])]
        if not variaveis.empty:
            top10 = variaveis.nlargest(10, "valor")[["descricao", "categoria", "cartao", "valor"]].copy()
            top10["valor"] = top10["valor"].apply(lambda x: f"R$ {x:,.2f}")
            top10.columns = ["Descrição", "Categoria", "Cartão", "Valor"]
            st.dataframe(top10, use_container_width=True, hide_index=True)
        else:
            st.info("Nenhum lançamento variável neste mês.")

        # ── Sprint 5: Painel de divisão (Kelvin/Thais) — não se aplica à Mãe ──
        if not _usuario_mae:
            st.divider()
            st.markdown("#### Divisão do mês")
            div_k = int(cfg.get("divisao_kelvin", 80))
            div_t = int(cfg.get("divisao_thais", 20))
            nome_k = cfg.get("nome_kelvin", "Kelvin")
            nome_t = cfg.get("nome_thais", "Thais")
            divisao = calcular_divisao_mes(mes, div_k, div_t)

            total_div = divisao["total"]
            k_paga = divisao["kelvin"]
            t_proporcional = divisao["thais_proporcional"]
            t_direto = divisao["thais_direto"]

            dk1, dk2, dk3, dk4 = st.columns(4)
            dk1.html(_card("Total fatura", f"R$ {total_div:,.2f}"))
            dk2.html(_card(f"{nome_k} paga ({div_k}%)", f"R$ {k_paga:,.2f}"))
            dk3.html(_card(f"{nome_t} — cota ({div_t}%)", f"R$ {t_proporcional:,.2f}"))
            dk4.html(_card(f"{nome_t} deve (direto)", f"R$ {t_direto:,.2f}",
                           nota="valor a ser ressarcido" if t_direto > 0 else None))

            if t_direto > 0:
                st.info(
                    f"💡 **{nome_t}** tem R$ {t_direto:,.2f} em lançamentos marcados como dela — "
                    f"valor a ressarcir para {nome_k}."
                )

            # Detalhamento dos lançamentos com Pessoa preenchida
            mask_p = lanc["pessoa_thais"].notna() & (lanc["valor_thais"].notna())
            if mask_p.any():
                df_pessoa = lanc[mask_p][["descricao", "categoria", "pessoa_thais", "valor_thais", "valor"]].copy()
                df_pessoa.columns = ["Descrição", "Categoria", "Pessoa", "Valor (Pessoa)", "Valor Total"]
                df_pessoa["Valor (Pessoa)"] = df_pessoa["Valor (Pessoa)"].apply(lambda x: f"R$ {float(x):,.2f}")
                df_pessoa["Valor Total"] = df_pessoa["Valor Total"].apply(lambda x: f"R$ {float(x):,.2f}")
                with st.expander(f"Ver detalhes dos {mask_p.sum()} lançamentos divididos"):
                    st.dataframe(df_pessoa, use_container_width=True, hide_index=True)

        # ── Sprint 5: Orçamento por categoria ────────────────────────────────
        st.divider()
        st.markdown("#### Orçamento por categoria")
        orcamentos = get_orcamentos(mes)
        por_cat_real = lanc.groupby("categoria")["valor"].sum().to_dict()

        # Configurar orçamentos (expander)
        with st.expander("⚙️ Configurar orçamentos do mês"):
            st.caption("Defina um teto de gastos para cada categoria. Deixe 0 para não monitorar.")
            orc_cols = st.columns(3)
            novos_orcs = {}
            for i, cat in enumerate(CATEGORIAS):
                col = orc_cols[i % 3]
                atual = float(orcamentos.get(cat, 0))
                novo = col.number_input(cat, min_value=0.0, step=50.0, value=atual,
                                        format="%.0f", key=f"orc_{cat}")
                novos_orcs[cat] = novo
            if st.button("💾", use_container_width=True, help="Salvar orçamentos"):
                for cat, val in novos_orcs.items():
                    if val > 0:
                        set_orcamento(mes, cat, val)
                    else:
                        delete_orcamento(mes, cat)
                st.success("Orçamentos atualizados!")
                st.rerun()

        # Barras de progresso por categoria
        cats_monitoradas = {c: v for c, v in orcamentos.items() if float(v) > 0}
        cats_sem_orcamento = [c for c in por_cat_real if c not in cats_monitoradas and float(por_cat_real[c]) > 0]

        if cats_monitoradas:
            prog_rows = ""
            for cat, limite in sorted(cats_monitoradas.items()):
                gasto = float(por_cat_real.get(cat, 0))
                limite = float(limite)
                pct = min(gasto / limite * 100, 100) if limite > 0 else 0
                over = gasto > limite
                bar_color = "#ef4444" if over else ("#f59e0b" if pct > 80 else "#4ade80")
                alerta = "⚠️" if over else ("🔶" if pct > 80 else "✅")
                prog_rows += f"""
                <tr style="border-bottom:1px solid rgba(255,255,255,0.06)">
                  <td style="padding:10px 14px;font-size:13px;width:160px">{alerta} {cat}</td>
                  <td style="padding:10px 14px;width:60%">
                    <div style="background:rgba(255,255,255,0.07);border-radius:4px;height:10px;overflow:hidden">
                      <div style="background:{bar_color};width:{pct:.1f}%;height:100%;border-radius:4px;transition:width .3s"></div>
                    </div>
                  </td>
                  <td style="padding:10px 14px;text-align:right;font-size:12px;color:#aaa;white-space:nowrap">
                    R$ {gasto:,.0f} / R$ {limite:,.0f}
                  </td>
                  <td style="padding:10px 14px;text-align:right;font-size:12px;font-weight:600;white-space:nowrap;color:{bar_color}">
                    {pct:.0f}%{'  (+R$ ' + f'{gasto-limite:,.0f})' if over else ''}
                  </td>
                </tr>"""
            st.html(f"""
            <table style="width:100%;border-collapse:collapse;font-family:inherit">
              <thead>
                <tr style="border-bottom:2px solid rgba(255,255,255,0.12)">
                  <th style="padding:8px 14px;text-align:left;font-size:11px;color:#666;text-transform:uppercase;letter-spacing:.6px">Categoria</th>
                  <th style="padding:8px 14px;text-align:left;font-size:11px;color:#666;text-transform:uppercase;letter-spacing:.6px">Progresso</th>
                  <th style="padding:8px 14px;text-align:right;font-size:11px;color:#666;text-transform:uppercase;letter-spacing:.6px">Gasto / Limite</th>
                  <th style="padding:8px 14px;text-align:right;font-size:11px;color:#666;text-transform:uppercase;letter-spacing:.6px">%</th>
                </tr>
              </thead>
              <tbody>{prog_rows}</tbody>
            </table>""")

            n_over = sum(1 for cat, lim in cats_monitoradas.items()
                         if float(por_cat_real.get(cat, 0)) > float(lim))
            if n_over:
                st.warning(f"⚠️ {n_over} categoria(s) acima do orçamento este mês.")
        else:
            st.caption("Nenhum orçamento configurado. Use o painel acima para definir limites por categoria.")

        if cats_sem_orcamento:
            nomes = ", ".join(cats_sem_orcamento)
            st.caption(f"Categorias sem orçamento com gastos: **{nomes}**")


# ══════════════════════════════════════════════════════════════════════════════
# HISTÓRICO
# ══════════════════════════════════════════════════════════════════════════════
elif pagina == "Histórico":
    st.title("Histórico")
    st.caption("Evolução de gastos ao longo dos meses disponíveis.")

    todos_meses = get_meses()
    if len(todos_meses) < 1:
        st.info("Nenhum mês com dados disponível.")
    else:
        labels_all = [MES_LABELS.get(m, m) for m in todos_meses]

        # Filtros de período
        fc1, fc2, fc3, fc4 = st.columns(4)
        idx_ini_def = 0
        idx_fim_def = len(todos_meses) - 1
        sel_ini = fc1.selectbox("De", labels_all, index=idx_ini_def, key="hist_ini")
        sel_fim = fc2.selectbox("Até", labels_all, index=idx_fim_def, key="hist_fim")
        filtro_cat_hist = fc3.multiselect("Categorias", CATEGORIAS, key="hist_cat")
        filtro_cartao_hist = fc4.multiselect("Cartão", CARTOES, key="hist_cart")

        idx_ini = labels_all.index(sel_ini)
        idx_fim = labels_all.index(sel_fim)
        if idx_ini > idx_fim:
            idx_ini, idx_fim = idx_fim, idx_ini
        meses_range = todos_meses[idx_ini: idx_fim + 1]

        # Carrega todos os lançamentos do período
        frames = []
        for m in meses_range:
            df_m = get_lancamentos(m)
            if not df_m.empty:
                df_m = df_m.copy()
                df_m["_mes"] = m
                frames.append(df_m)

        if not frames:
            st.info("Nenhum lançamento no período selecionado.")
        else:
            df_hist = pd.concat(frames, ignore_index=True)
            if filtro_cat_hist:
                df_hist = df_hist[df_hist["categoria"].isin(filtro_cat_hist)]
            if filtro_cartao_hist:
                df_hist = df_hist[df_hist["cartao"].isin(filtro_cartao_hist)]

            df_hist["Mês"] = df_hist["_mes"].map(lambda x: MES_LABELS.get(x, x))

            # ── Gráfico de linha: evolução do total mensal ────────────────────
            st.markdown("#### Evolução dos gastos totais")
            totais = df_hist.groupby(["_mes", "Mês"], sort=False)["valor"].sum().reset_index()
            totais = totais.sort_values("_mes")
            fig_line = px.line(totais, x="Mês", y="valor", markers=True,
                labels={"valor": "Total (R$)"},
                color_discrete_sequence=["#4C9BE8"])
            fig_line.update_traces(line_width=2.5, marker_size=8,
                text=totais["valor"].apply(lambda v: f"R$ {v:,.0f}"),
                textposition="top center")
            fig_line.update_layout(height=300, margin=dict(t=20, b=20),
                yaxis_title="Total (R$)", xaxis_title="")
            st.plotly_chart(fig_line, use_container_width=True)

            # ── Gráfico de barras empilhadas: categorias por mês ─────────────
            st.markdown("#### Gastos por categoria por mês")
            cat_mes = df_hist.groupby(["_mes", "Mês", "categoria"], sort=False)["valor"].sum().reset_index()
            cat_mes = cat_mes.sort_values("_mes")
            fig_stack = px.bar(cat_mes, x="Mês", y="valor", color="categoria",
                barmode="stack", labels={"valor": "Total (R$)", "categoria": "Categoria"},
                color_discrete_sequence=px.colors.qualitative.Set2)
            fig_stack.update_layout(height=350, margin=dict(t=20, b=20),
                yaxis_title="Total (R$)", xaxis_title="", legend_title="Categoria")
            st.plotly_chart(fig_stack, use_container_width=True)

            st.divider()
            col_l, col_r = st.columns(2)

            # ── Top 10 maiores gastos do período ──────────────────────────────
            with col_l:
                st.markdown("#### Top 10 maiores gastos (variáveis)")
                var_hist = df_hist[~df_hist["tipo_parcela"].isin(["FIXO"])]
                if not var_hist.empty:
                    top10 = var_hist.nlargest(10, "valor")[["descricao", "Mês", "categoria", "valor"]].copy()
                    top10["valor"] = top10["valor"].apply(lambda x: f"R$ {x:,.2f}")
                    top10.columns = ["Descrição", "Mês", "Categoria", "Valor"]
                    st.dataframe(top10, use_container_width=True, hide_index=True)
                else:
                    st.info("Nenhum gasto variável no período.")

            # ── Comparativo por categoria ─────────────────────────────────────
            with col_r:
                st.markdown("#### Comparativo por categoria")
                comp = df_hist.groupby("categoria")["valor"].agg(["sum", "mean", "count"]).reset_index()
                comp.columns = ["Categoria", "Total (R$)", "Média/mês (R$)", "Lançamentos"]
                comp["Total (R$)"] = comp["Total (R$)"].round(2)
                comp["Média/mês (R$)"] = comp["Média/mês (R$)"].round(2)
                comp = comp.sort_values("Total (R$)", ascending=False)
                st.dataframe(comp, use_container_width=True, hide_index=True)

            # ── Resumo do período ─────────────────────────────────────────────
            st.divider()
            st.markdown("#### Resumo do período")
            total_periodo = df_hist["valor"].sum()
            media_mensal = total_periodo / len(meses_range) if meses_range else 0
            mes_maior = totais.loc[totais["valor"].idxmax(), "Mês"] if not totais.empty else "—"
            mes_menor = totais.loc[totais["valor"].idxmin(), "Mês"] if not totais.empty else "—"

            def _card(label, value, nota=None, vermelho=False):
                cor = "#ff4b4b" if vermelho else "#21c354"
                nota_html = f'<div style="font-size:12px;color:{cor};margin-top:4px">{nota}</div>' if nota else ""
                return (f'<div style="padding:6px 0">'
                        f'<div style="font-size:13px;color:#aaa;margin-bottom:4px">{label}</div>'
                        f'<div style="font-size:1.4rem;font-weight:700">{value}</div>'
                        f'{nota_html}</div>')

            rs1, rs2, rs3, rs4 = st.columns(4)
            rs1.html(_card("Total no período", f"R$ {total_periodo:,.0f}"))
            rs2.html(_card("Média mensal", f"R$ {media_mensal:,.0f}"))
            rs3.html(_card("Mês mais caro", mes_maior))
            rs4.html(_card("Mês mais barato", mes_menor))


# ══════════════════════════════════════════════════════════════════════════════
# LANÇAMENTOS
# ══════════════════════════════════════════════════════════════════════════════
elif pagina == "Lançamentos":
    # Cada cartão com sub-bandeira tem sua própria key de filtro (multiselect).
    SUBTIPO_FILTRO_KEY = {"Santander": "lanc_filtro_subtipo", "Itaú": "lanc_filtro_subtipo_itau"}

    # Resolve cliques nos cards de "Totais por cartão" ANTES de criar os widgets de
    # filtro (a key de um widget não pode ser escrita depois de instanciado no mesmo run).
    # Clicar num card SUBSTITUI o filtro pelo card clicado (não acumula); clicar de novo
    # no card já ativo limpa o filtro.
    _toggle = st.session_state.pop("_toggle_cartao_request", None)
    if _toggle is not None:
        c_tog, sub_tog = _toggle
        sub_key = SUBTIPO_FILTRO_KEY.get(c_tog)
        subtipo_atual_do_clicado = list(st.session_state.get(sub_key, [])) if sub_key else []
        ja_era_unico_selecionado = (
            list(st.session_state.get("lanc_filtro_cartao", [])) == [c_tog]
            and ((sub_tog is None and not subtipo_atual_do_clicado)
                 or (sub_tog is not None and subtipo_atual_do_clicado == [sub_tog]))
        )
        if ja_era_unico_selecionado:
            st.session_state["lanc_filtro_cartao"] = []
            st.session_state["lanc_filtro_subtipo"] = []
            st.session_state["lanc_filtro_subtipo_itau"] = []
        else:
            st.session_state["lanc_filtro_cartao"] = [c_tog]
            st.session_state["lanc_filtro_subtipo"] = [sub_tog] if sub_key == "lanc_filtro_subtipo" else []
            st.session_state["lanc_filtro_subtipo_itau"] = [sub_tog] if sub_key == "lanc_filtro_subtipo_itau" else []

    col_title, col_fixos = st.columns([5, 1])
    col_title.title(f"Lançamentos — {mes_label}")
    with col_fixos:
        st.write("")
        if st.button("📌", use_container_width=True, help="Aplicar fixos do mês"):
            n = aplicar_fixos_ao_mes(mes)
            st.success(f"{n} fixo(s) adicionado(s)!" if n > 0 else "Todos os fixos já estão neste mês.")
            # Sem st.rerun(): os widgets de filtro só são criados depois, mais abaixo no
            # script — rerodar aqui descartaria o estado deles (filtro "sumia" ao salvar).
            # O restante do script já reflete os dados atualizados nesta mesma execução.

    if "editando_id" not in st.session_state:
        st.session_state.editando_id = None
    if "form_novo_ver" not in st.session_state:
        st.session_state.form_novo_ver = 0

    def form_lancamento(prefixo, dados=None):
        """Renderiza campos do formulário. Funciona fora de st.form para reatividade."""
        d = dados or {}
        # Editando uma parcela que já faz parte de um parcelamento em andamento (tem
        # id_grupo) — nesse caso "Valor total" é o valor DESTA parcela, não da compra
        # inteira (diferente do fluxo de criação, que sempre divide pelo nº de parcelas).
        ja_era_parcela_existente = (
            dados is not None
            and d.get("tipo_parcela") in ("parcelado", "ULTIMA")
            and d.get("id_grupo") and not pd.isna(d.get("id_grupo"))
        )
        fc1, fc2 = st.columns(2)
        cartao_atual = d.get("cartao", "Santander")
        cartao_idx = CARTOES.index(cartao_atual) if cartao_atual in CARTOES else 0
        cartao = fc1.selectbox("Cartão", CARTOES, index=cartao_idx, key=f"{prefixo}_cartao")
        valor = fc2.number_input("Valor total (R$)", step=0.01, format="%.2f",
            value=float(d.get("valor") or 0.01), key=f"{prefixo}_valor",
            help="Negativo = crédito (aparece em verde na tabela)")

        subtipo = None
        if cartao == "Santander":
            sub_atual = d.get("subtipo_cartao") or "Virtual"
            sub_idx = SUBTIPOS_SANTANDER.index(sub_atual) if sub_atual in SUBTIPOS_SANTANDER else 0
            subtipo = st.radio("Tipo Santander", SUBTIPOS_SANTANDER, index=sub_idx,
                horizontal=True, key=f"{prefixo}_subtipo")
        elif cartao == "Itaú":
            sub_atual = d.get("subtipo_cartao") or "Visa"
            sub_idx = SUBTIPOS_ITAU.index(sub_atual) if sub_atual in SUBTIPOS_ITAU else 0
            subtipo = st.radio("Bandeira Itaú", SUBTIPOS_ITAU, index=sub_idx,
                horizontal=True, key=f"{prefixo}_subtipo_itau")

        descricao = st.text_input("Descrição", value=d.get("descricao", ""), key=f"{prefixo}_desc")
        fc3, fc4 = st.columns(2)
        cat_idx = CATEGORIAS.index(d["categoria"]) if d.get("categoria") in CATEGORIAS else 0
        categoria = fc3.selectbox("Categoria", CATEGORIAS, index=cat_idx, key=f"{prefixo}_cat")
        tipos = ["única", "FIXO", "ULTIMA", "parcelado"]
        tipo_idx = tipos.index(d["tipo_parcela"]) if d.get("tipo_parcela") in tipos else 0
        tipo = fc4.selectbox("Tipo", tipos, index=tipo_idx, key=f"{prefixo}_tipo")

        # Dividir — checkbox reativo (sem st.form)
        tem_pessoa = bool(d.get("pessoa_thais") and not pd.isna(d.get("pessoa_thais", "")))
        dividir = st.checkbox("Dividir", value=tem_pessoa, key=f"{prefixo}_dividir")

        pessoa = None
        val_pessoa = None
        if dividir:
            col_p1, col_p2 = st.columns(2)
            pessoa_default = d.get("pessoa_thais") or "Thais"
            if pd.isna(pessoa_default):
                pessoa_default = "Thais"
            pessoa = col_p1.text_input("Pessoa", value=str(pessoa_default), key=f"{prefixo}_pessoa")

            val_key = f"{prefixo}_valpessoa"
            auto_key = f"{prefixo}_valpessoa_auto"
            div_t = int(get_config().get("divisao_thais", 20))

            if str(pessoa).strip().lower() == "thais":
                new_auto = round(valor * div_t / 100, 2)
                prev_auto = st.session_state.get(auto_key)
                curr_val = st.session_state.get(val_key)
                # Atualiza automaticamente se: primeira vez ou usuário não mudou manualmente
                if curr_val is None or curr_val == prev_auto:
                    st.session_state[val_key] = new_auto
                st.session_state[auto_key] = new_auto
            else:
                st.session_state.pop(auto_key, None)
                if val_key not in st.session_state:
                    vp_default = float(d.get("valor_thais") or 0.0)
                    st.session_state[val_key] = 0.0 if pd.isna(vp_default) else vp_default

            # Sem `value=`: a key já garante o valor inicial acima, e passar os dois juntos
            # gera o aviso do Streamlit "created with a default value but also had its value
            # set via the Session State API".
            val_pessoa = col_p2.number_input("Valor (R$)", step=0.01,
                format="%.2f", key=val_key)

        total_parc = None
        mes_inicio_parc = mes
        if tipo == "parcelado":
            if ja_era_parcela_existente:
                total_parc = st.number_input("Total de parcelas (faltam após esta)", min_value=1,
                    max_value=48, value=int(d.get("total_parcelas") or 1), step=1,
                    key=f"{prefixo}_parc",
                    help="Parcelas restantes depois desta (ex.: 4 = mais 4 parcelas após esta).")
                st.caption("✏️ Editando uma parcela já existente — o **valor acima é o valor "
                           "desta parcela** (não da compra inteira). Ao salvar, atualiza também "
                           "as próximas parcelas deste mesmo parcelamento.")
            else:
                modo_valor = st.radio("O valor informado acima é", ["Valor total da compra", "Valor de cada parcela"],
                    horizontal=True, key=f"{prefixo}_modo_valor",
                    help="\"Valor total\" divide pelo nº de parcelas. \"Valor de cada parcela\" "
                         "usa o número digitado direto, sem dividir.")
                pp1, pp2 = st.columns(2)
                total_parc = pp1.number_input("Total de parcelas", min_value=1, max_value=48,
                    value=int(d.get("total_parcelas") or 2), step=1, key=f"{prefixo}_parc")
                meses_disp = get_meses()
                idx_mes = meses_disp.index(mes) if mes in meses_disp else 0
                labels_m = [MES_LABELS.get(m, m) for m in meses_disp]
                escolha_m = pp2.selectbox("Mês da 1ª parcela", labels_m, index=idx_mes, key=f"{prefixo}_mesinicio")
                mes_inicio_parc = meses_disp[labels_m.index(escolha_m)]
                if total_parc:
                    if modo_valor == "Valor total da compra":
                        vparc = valor / int(total_parc)
                        st.caption(f"💳 Informe o **valor total** acima — serão **{int(total_parc)}x de "
                                   f"R\\$ {vparc:,.2f}** (total R\\$ {valor:,.2f}).")
                    else:
                        st.caption(f"💳 Cada parcela será de **R\\$ {valor:,.2f}** — "
                                   f"**{int(total_parc)}x** (total R\\$ {valor * int(total_parc):,.2f}).")

        return {
            "cartao": cartao, "subtipo": subtipo, "valor": valor, "descricao": descricao,
            "categoria": categoria, "tipo": tipo,
            "pessoa": pessoa.strip() if pessoa and str(pessoa).strip() else None,
            "val_pessoa": val_pessoa if dividir and val_pessoa else None,
            "total_parc": total_parc,
            "mes_inicio_parc": mes_inicio_parc,
            "propagar_grupo": ja_era_parcela_existente,
            "valor_eh_parcela": (tipo == "parcelado" and not ja_era_parcela_existente
                                  and st.session_state.get(f"{prefixo}_modo_valor") == "Valor de cada parcela"),
        }

    def salvar_campos(campos):
        if not campos["descricao"]:
            st.error("Preencha a descrição.")
            return False
        if campos["tipo"] == "parcelado" and campos["total_parc"]:
            n_parc = int(campos["total_parc"])
            if campos.get("valor_eh_parcela"):
                # O valor informado já é o de CADA parcela → usa direto, sem dividir
                valor_parc = round(campos["valor"], 2)
                vt_parc = round(campos["val_pessoa"], 2) if campos["val_pessoa"] else None
            else:
                # O valor informado é o TOTAL da compra → divide pelo nº de parcelas
                valor_parc = round(campos["valor"] / n_parc, 2)
                vt_parc = round(campos["val_pessoa"] / n_parc, 2) if campos["val_pessoa"] else None
            gid = criar_grupo_parcelamento(
                descricao=campos["descricao"], cartao=campos["cartao"],
                subtipo_cartao=campos["subtipo"], categoria=campos["categoria"],
                valor_parcela=valor_parc, total_parcelas=n_parc,
                mes_inicio=campos["mes_inicio_parc"],
                pessoa_thais=campos["pessoa"], valor_thais=vt_parc,
            )
            ultimo = _proximo_mes(campos["mes_inicio_parc"], n_parc - 1)
            st.success(f"'{campos['descricao']}' criado em {n_parc}x de R$ {valor_parc:,.2f} "
                       f"(total R$ {valor_parc * n_parc:,.2f}) até {ultimo}!")
        else:
            add_lancamento(
                mes_ano=mes, cartao=campos["cartao"], dono="Kelvin",
                valor=campos["valor"], descricao=campos["descricao"],
                categoria=campos["categoria"],
                valor_thais=campos["val_pessoa"], pessoa_thais=campos["pessoa"],
                tipo_parcela=campos["tipo"], total_parcelas=campos["total_parc"],
                subtipo_cartao=campos["subtipo"],
            )
            st.success(f"Lançamento '{campos['descricao']}' salvo!")
        return True

    # ── Filtros ──────────────────────────────────────────────────────────────
    # (vem ANTES do "Novo lançamento" propositalmente: o botão Salvar chama
    # st.rerun() e, se os filtros fossem instanciados só depois, o Streamlit
    # descartaria o estado deles nesse ciclo abortado — fazendo o filtro "sumir"
    # toda vez que um lançamento era salvo.)
    st.markdown("#### Filtrar")
    fc1, fc2, fc3, fc4, fc5, fc6 = st.columns([2, 2, 1.3, 1.3, 1.3, 1.3])
    # Limpa valores obsoletos do session_state antes de renderizar os multiselects
    for _k, _opts in [("lanc_filtro_subtipo", SUBTIPOS_SANTANDER),
                      ("lanc_filtro_subtipo_itau", SUBTIPOS_ITAU)]:
        if _k in st.session_state:
            st.session_state[_k] = [v for v in st.session_state[_k] if v in _opts]
    filtro_cat = fc1.multiselect("Categoria", CATEGORIAS, key="lanc_filtro_cat")
    filtro_tipo = fc2.multiselect("Tipo", ["única", "FIXO", "ULTIMA", "parcelado"], key="lanc_filtro_tipo")
    filtro_cartao = fc3.multiselect("Cartão", CARTOES, key="lanc_filtro_cartao")
    filtro_subtipo = fc4.multiselect("Santander", SUBTIPOS_SANTANDER, key="lanc_filtro_subtipo")
    filtro_subtipo_itau = fc5.multiselect("Itaú", SUBTIPOS_ITAU, key="lanc_filtro_subtipo_itau")
    filtro_ordem = fc6.selectbox("Ordenar por", ["Mais antigos", "Mais recentes"], key="lanc_filtro_ordem")
    busca = st.text_input("Buscar descrição", placeholder="Ex: Spotify, Uber...", key="lanc_busca")

    if get_usuario_atual() == "mae":
        with st.expander("📋 Fixas (orçamento mensal)", expanded=False):
            _editor_controle_extra("fixas", mes, "fixas")
        with st.expander("🍽️ Restaurante", expanded=False):
            _editor_controle_extra("restaurante", mes, "restaurante", label_nome="Dia")
        with st.expander("🏠 Aluguel", expanded=False):
            _editor_controle_extra("aluguel", mes, "aluguel", label_nome="Pessoa", mostrar_nota=True,
                                    ajuda="Copiado automaticamente do mês anterior ao criar um mês novo.")

    # ── Budget provisório Santander (Combustível / Feira) ─────────────────────
    _cfg = get_config()
    _bud_comb  = float(_cfg.get("budget_combustivel", 700))
    _bud_feira = float(_cfg.get("budget_feira", 200))
    _zerado    = _cfg.get(f"budget_zerado_{mes}", "0") == "1"

    _lanc_full_bud = get_lancamentos(mes)
    _gasto_comb  = 0.0
    _gasto_feira = 0.0
    if not _lanc_full_bud.empty:
        _sant = _lanc_full_bud[_lanc_full_bud["cartao"] == "Santander"]
        _gasto_comb  = float(_sant[_sant["descricao"].str.contains("combustiv", case=False, na=False)]["valor"].sum())
        _gasto_feira = float(_sant[_sant["descricao"].str.contains("feira", case=False, na=False)]["valor"].sum())

    _rest_comb  = max(0.0, _bud_comb  - _gasto_comb)  if not _zerado else 0.0
    _rest_feira = max(0.0, _bud_feira - _gasto_feira) if not _zerado else 0.0

    st.markdown("##### ⛽ Budget Santander — Combustível & Feira")
    bc1, bc2, bc3 = st.columns([2, 2, 1])
    with bc1:
        novo_bud_comb = st.number_input("⛽ Combustível — orçamento (R$)", value=_bud_comb,
                                        min_value=0.0, step=50.0, format="%.0f", key="bud_comb")
        pct_c = int(min(_gasto_comb / _bud_comb * 100, 100)) if _bud_comb > 0 else 0
        st.progress(pct_c / 100)
        def _brl(v): return f"R$ {v:,.2f}".replace(",","X").replace(".",",").replace("X",".")
        st.html(f"<small style='color:#aaa'>Gasto: {_brl(_gasto_comb)} &nbsp;·&nbsp; Restante: <b style='color:#fff'>{_brl(_rest_comb)}</b></small>")
    with bc2:
        novo_bud_feira = st.number_input("🛒 Feira — orçamento (R$)", value=_bud_feira,
                                         min_value=0.0, step=50.0, format="%.0f", key="bud_feira")
        pct_f = int(min(_gasto_feira / _bud_feira * 100, 100)) if _bud_feira > 0 else 0
        st.progress(pct_f / 100)
        st.html(f"<small style='color:#aaa'>Gasto: {_brl(_gasto_feira)} &nbsp;·&nbsp; Restante: <b style='color:#fff'>{_brl(_rest_feira)}</b></small>")
    with bc3:
        st.write(""); st.write(""); st.write("")
        if st.button("💾", key="bud_salvar", help="Salvar orçamentos", use_container_width=True):
            set_config("budget_combustivel", novo_bud_comb)
            set_config("budget_feira", novo_bud_feira)
            st.success("Salvo!")
            st.rerun()
        _label_zero = "✓" if _zerado else "🔄"
        if st.button(_label_zero, key="bud_zerar", use_container_width=True,
                     help="Zerar provisões do mês" if not _zerado else "Restaurar provisões"):
            set_config(f"budget_zerado_{mes}", "0" if _zerado else "1")
            st.rerun()
    st.divider()

    # ── Novo lançamento (sem st.form → checkbox reativo) ──────────────────────
    ver = st.session_state.form_novo_ver
    with st.expander("➕ Novo lançamento", expanded=False):
        campos = form_lancamento(f"novo_{ver}")
        if st.button("💾 Salvar", use_container_width=True, key=f"btn_novo_{ver}"):
            if salvar_campos(campos):
                st.session_state.form_novo_ver = ver + 1
                st.rerun()

    lanc = get_lancamentos(mes)
    if "subtipo_cartao" not in lanc.columns:
        lanc["subtipo_cartao"] = None

    if not lanc.empty:
        if filtro_cat:
            lanc = lanc[lanc["categoria"].isin(filtro_cat)]
        if filtro_tipo:
            lanc = lanc[lanc["tipo_parcela"].isin(filtro_tipo)]
        if filtro_cartao:
            lanc = lanc[lanc["cartao"].isin(filtro_cartao)]
        if filtro_subtipo:
            lanc = lanc[(lanc["cartao"] != "Santander") | lanc["subtipo_cartao"].isin(filtro_subtipo)]
        if filtro_subtipo_itau:
            lanc = lanc[(lanc["cartao"] != "Itaú") | lanc["subtipo_cartao"].isin(filtro_subtipo_itau)]
        if busca:
            lanc = lanc[lanc["descricao"].str.contains(busca, case=False, na=False)]

    # Ordenação ESTÁVEL por id (evita reorganização ao marcar/desmarcar checkbox).
    # id crescente = mais antigos primeiro, mais novos sempre no final.
    if not lanc.empty:
        lanc = lanc.sort_values("id", ascending=(filtro_ordem == "Mais antigos"))

    # ── Tabela de lançamentos ─────────────────────────────────────────────────
    if lanc.empty:
        st.info("Nenhum lançamento encontrado para os filtros selecionados.")
    else:
        # ── Totais por cartão (item 2) — usa o mês completo para conferência ──
        lanc_mes_full = get_lancamentos(mes)
        tot_cartao = {}
        for _, r in lanc_mes_full.iterrows():
            c = str(r["cartao"])
            sub = r.get("subtipo_cartao")
            sub_str = str(sub) if sub and not pd.isna(sub) else None
            if c == "Santander":
                sub_s = "Físico" if sub_str == "Físico" else None
            elif c == "Itaú":
                sub_s = sub_str if sub_str in SUBTIPOS_ITAU else None
            else:
                sub_s = None
            chave = (c, sub_s)
            tot_cartao[chave] = tot_cartao.get(chave, 0.0) + float(r["valor"])
        # Adiciona provisão restante ao total Santander
        _prov = _rest_comb + _rest_feira
        if _prov > 0:
            tot_cartao[("Santander", None)] = tot_cartao.get(("Santander", None), 0.0) + _prov
        st.markdown("##### Totais por cartão no mês &nbsp;<small style='color:#666;font-weight:400'>(clique para filtrar)</small>", unsafe_allow_html=True)
        cartoes_ordenados = sorted(tot_cartao.items(), key=lambda x: -x[1])
        sub_filtro_ativo = {"Santander": filtro_subtipo, "Itaú": filtro_subtipo_itau}

        # Cards + botão de filtro juntos em colunas iguais (alinhamento garantido)
        cols_cards = st.columns(len(cartoes_ordenados) if cartoes_ordenados else 1)
        for i, ((c, sub_s), tot) in enumerate(cartoes_ordenados):
            ativo = (c in filtro_cartao) and (sub_s is None or sub_s in sub_filtro_ativo.get(c, []))
            label_banco = f"{c} {sub_s}" if sub_s else c
            with cols_cards[i]:
                st.html(card_cartao(c, tot, sub_s, ativo=ativo))
                clicado = st.button(
                    "✓" if ativo else "◎",
                    key=f"cardbtn_cartao_{i}",
                    use_container_width=True,
                    type="primary" if ativo else "secondary",
                    help=label_banco,
                )
            if clicado:
                st.session_state["_toggle_cartao_request"] = (c, sub_s)
                st.rerun()

        st.markdown(f"<small style='color:#888'>{len(lanc)} lançamentos exibidos · Total filtrado: <b>R$ {lanc['valor'].sum():,.2f}</b></small>", unsafe_allow_html=True)

        def _toggle_conferido(lid, key):
            update_lancamento(lid, conferido=bool(st.session_state[key]))

        # Cabeçalho
        hc = st.columns([0.5, 6.5, 0.6, 0.6])
        hc[0].html('<div style="color:#666;font-size:10px;text-transform:uppercase;letter-spacing:.4px;'
                   'font-weight:500;border-bottom:2px solid rgba(255,255,255,0.12);padding:6px 0;text-align:center">✓</div>')
        hc[1].html(lancamento_header())

        # Linhas com checkbox Conferido + botões de ação por registro
        for _, row in lanc.iterrows():
            lid = int(row["id"])
            conf_atual = bool(row.get("conferido")) if not pd.isna(row.get("conferido")) else False
            subtipo_val = row.get("subtipo_cartao")
            subtipo_str = str(subtipo_val) if subtipo_val and not pd.isna(subtipo_val) else None
            pessoa_val = row.get("pessoa_thais")
            nome_p = str(pessoa_val) if pessoa_val and not pd.isna(pessoa_val) else None
            valor_p = row.get("valor_thais")
            val_p = float(valor_p) if valor_p and not pd.isna(valor_p) else None
            tipo_raw = str(row["tipo_parcela"])
            parc_rest = row.get("total_parcelas")
            if tipo_raw == "ULTIMA":
                faltam = "ÚLTIMA"
            elif tipo_raw == "parcelado" and parc_rest is not None and not pd.isna(parc_rest):
                faltam = str(int(parc_rest))
            else:
                faltam = "—"
            rc = st.columns([0.5, 6.5, 0.6, 0.6])
            rc[0].checkbox("Conferido", value=conf_atual, key=f"conf_{lid}",
                           on_change=_toggle_conferido, args=(lid, f"conf_{lid}"),
                           label_visibility="collapsed")
            rc[1].html(lancamento_row(
                descricao=_fmt_desc(row["descricao"]),
                cartao=str(row["cartao"]),
                valor=float(row["valor"]),
                categoria=str(row.get("categoria", "")),
                tipo=tipo_raw,
                faltam=faltam,
                subtipo=subtipo_str,
                conferido=conf_atual,
                pessoa=nome_p,
                valor_pessoa=val_p,
            ))
            if rc[2].button("✏️", key=f"edit_{lid}", help="Editar"):
                st.session_state.editando_id = lid
                st.rerun()
            if rc[3].button("🗑", key=f"del_{lid}", help="Excluir"):
                delete_lancamento(lid)
                st.rerun()

        # Contador de conferência
        n_conf = int(lanc["conferido"].fillna(False).astype(bool).sum()) if "conferido" in lanc.columns else 0
        st.caption(f"✅ {n_conf} de {len(lanc)} lançamentos conferidos.")

        # ── Formulário de edição (popup, não precisa rolar a página) ─────────
        @st.dialog("✏️ Editar lançamento")
        def _dialog_editar(lid_ed, descricao_ed, dados_ed):
            st.caption(descricao_ed)
            edit_ver = st.session_state.get(f"edit_ver_{lid_ed}", 0)
            campos_ed = form_lancamento(f"edit_{lid_ed}_{edit_ver}", dados=dados_ed)
            ce1, ce2 = st.columns(2)
            if ce1.button("💾 Salvar", use_container_width=True, key="btn_salvar_ed"):
                update_lancamento(lid_ed,
                    cartao=campos_ed["cartao"], valor=campos_ed["valor"],
                    descricao=campos_ed["descricao"], categoria=campos_ed["categoria"],
                    tipo_parcela=campos_ed["tipo"],
                    pessoa_thais=campos_ed["pessoa"], valor_thais=campos_ed["val_pessoa"],
                    total_parcelas=campos_ed["total_parc"],
                    subtipo_cartao=campos_ed["subtipo"],
                )
                id_grupo = dados_ed.get("id_grupo")
                if campos_ed.get("propagar_grupo") and id_grupo and not pd.isna(id_grupo):
                    n_prop = propagar_parcela_grupo(
                        int(id_grupo), dados_ed.get("mes_ano"),
                        valor=campos_ed["valor"], categoria=campos_ed["categoria"],
                        pessoa_thais=campos_ed["pessoa"], valor_thais=campos_ed["val_pessoa"],
                    )
                    if n_prop > 1:
                        st.toast(f"Valor atualizado também nas próximas {n_prop - 1} parcela(s) deste parcelamento.")
                st.session_state.editando_id = None
                st.rerun()
            if ce2.button("✕ Cancelar", use_container_width=True, key="btn_cancel_ed"):
                st.session_state.editando_id = None
                st.rerun()

        if st.session_state.editando_id:
            row_ed = lanc[lanc["id"] == st.session_state.editando_id]
            if not row_ed.empty:
                row_ed = row_ed.iloc[0]
                _dialog_editar(st.session_state.editando_id, f"*{row_ed['descricao']}*", row_ed.to_dict())
            else:
                st.session_state.editando_id = None

        # ── Resumo por categoria ──────────────────────────────────────────────
        st.divider()
        st.markdown("#### Resumo por categoria")
        por_cat = lanc.groupby("categoria")["valor"].agg(["sum", "count"]).reset_index()
        por_cat.columns = ["Categoria", "Total (R$)", "Qtd"]
        por_cat["Total (R$)"] = por_cat["Total (R$)"].round(2)
        st.dataframe(por_cat.sort_values("Total (R$)", ascending=False), use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# PARCELAMENTOS
# ══════════════════════════════════════════════════════════════════════════════
elif pagina == "Parcelamentos":
    st.title("Parcelamentos em aberto")
    st.caption("Compras parceladas criadas pelo sistema. Parcelas futuras são geradas automaticamente.")

    grupos = get_grupos_ativos()
    df_lanc_all = load_sheet("lancamentos")
    if "subtipo_cartao" not in df_lanc_all.columns:
        df_lanc_all["subtipo_cartao"] = None

    if grupos.empty:
        st.info("Nenhum parcelamento ativo. Crie um lançamento do tipo 'parcelado' na aba Lançamentos.")
    else:
        ativos = grupos[grupos.get("restantes", 0) > 0] if "restantes" in grupos.columns else grupos
        st.markdown(f"**{len(ativos)} grupo(s) ativo(s)**")
        st.divider()

        class _MesShort:
            def get(self, ma, default=None):
                return fmt_mes(ma, curto=True) if ma else default
        MES_SHORT = _MesShort()

        for _, g in ativos.iterrows():
            gid = int(g["id"])
            total = int(g["total_parcelas"])
            pagas = int(g.get("pagas", 0))
            restantes = int(g.get("restantes", total - pagas))
            pct = int(pagas / total * 100) if total > 0 else 0
            valor_parc = float(g["valor_parcela"])

            sub = g.get("subtipo_cartao")
            sub_str = str(sub) if sub and not pd.isna(sub) else None
            badge = bank_badge(str(g["cartao"]), sub_str)

            col_info, col_prog, col_acao = st.columns([2.5, 2, 1.8])
            with col_info:
                st.html(
                    f'<div style="padding:4px 0">'
                    f'<span style="font-weight:600;font-size:15px">{g["descricao"]}</span> {badge}'
                    f'<br><span style="font-size:12px;color:#888">{g["categoria"]} · '
                    f'R$ {valor_parc:,.2f}/mês · total R$ {valor_parc*total:,.2f}</span></div>'
                )
            with col_prog:
                st.markdown(f"**{pagas}/{total}** parcelas pagas &nbsp; `{pct}%`")
                st.progress(pct / 100)
                prox_mes = _proximo_mes(str(g["mes_inicio"]), pagas)
                prox_label = MES_SHORT.get(prox_mes, prox_mes)
                if restantes == 1:
                    st.markdown(f"🔔 **Última parcela:** {prox_label}")
                elif restantes > 0:
                    st.markdown(f"Próxima: **{prox_label}** · faltam R$ {valor_parc*restantes:,.2f}")
            with col_acao:
                st.write("")
                if restantes > 0:
                    if st.button("✅", key=f"quit_{gid}", help="Quitar antecipado", use_container_width=True):
                        cancelar_parcelas_restantes(gid, mes)
                        st.success("Parcelas futuras removidas!")
                        st.rerun()
            st.divider()

    st.markdown("#### Projeção de comprometimento por mês")
    df_parc = df_lanc_all[df_lanc_all["tipo_parcela"].isin(["parcelado", "ULTIMA"])]
    if not df_parc.empty and "id_grupo" in df_parc.columns:
        df_pg = df_parc[df_parc["id_grupo"].notna()]
        if not df_pg.empty:
            proj = df_pg.groupby("mes_ano")["valor"].sum().reset_index()
            proj.columns = ["Mês", "Total parcelas (R$)"]
            proj["Total parcelas (R$)"] = proj["Total parcelas (R$)"].round(2)
            proj["Mês"] = proj["Mês"].map(lambda x: MES_LABELS.get(x, x))
            st.dataframe(proj.sort_values("Mês"), use_container_width=True, hide_index=True)
        else:
            st.info("Nenhum parcelamento criado via sistema ainda.")
    else:
        st.info("Nenhum parcelamento criado via sistema ainda.")


# ══════════════════════════════════════════════════════════════════════════════
# FIXOS
# ══════════════════════════════════════════════════════════════════════════════
elif pagina == "Fixos":
    st.title("Gastos Fixos")
    st.caption("Lançamentos que se repetem todo mês. Edite os valores estimados aqui.")

    fixos = get_fixos(apenas_ativos=False)

    with st.expander("➕ Novo fixo", expanded=False):
        # Sem st.form: precisa ser reativo para mostrar a bandeira/tipo certo
        # conforme o cartão escolhido (igual ao formulário de Lançamentos).
        ff1, ff2 = st.columns(2)
        f_cartao = ff1.selectbox("Cartão", CARTOES, key="f_cartao")
        f_desc = ff2.text_input("Descrição", key="f_desc")
        f_subtipo = None
        if f_cartao == "Santander":
            f_subtipo = st.radio("Tipo Santander", SUBTIPOS_SANTANDER, horizontal=True, key="f_subtipo_santander")
        elif f_cartao == "Itaú":
            f_subtipo = st.radio("Bandeira Itaú", SUBTIPOS_ITAU, horizontal=True, key="f_subtipo_itau")
        ff3, ff4 = st.columns(2)
        f_cat = ff3.selectbox("Categoria", CATEGORIAS, key="f_cat")
        f_val = ff4.number_input("Valor estimado (R$)", min_value=0.0, step=0.01, format="%.2f", key="f_val")
        ff5, ff6, ff7 = st.columns(3)
        f_divide = ff5.checkbox("Dividir?", key="f_divide")
        f_pessoa = ff6.text_input("Pessoa", value="Thais", key="f_pessoa")
        f_vt = ff7.number_input("Valor (R$)", min_value=0.0, step=0.01, format="%.2f", key="f_vt")
        if st.button("💾", use_container_width=True, key="f_salvar", help="Salvar fixo"):
            df_fixos = load_sheet("fixos")
            novo_id = int(df_fixos["id"].max() + 1) if not df_fixos.empty else 1
            novo = {
                "id": novo_id, "cartao": f_cartao, "subtipo_cartao": f_subtipo,
                "descricao": f_desc, "categoria": f_cat, "valor_estimado": f_val,
                "pessoa_thais": f_pessoa if f_divide else None,
                "valor_thais": f_vt if f_divide else None,
                "ativo": True,
            }
            df_fixos = pd.concat([df_fixos, pd.DataFrame([novo])], ignore_index=True)
            save_sheet("fixos", df_fixos)
            st.success(f"Fixo '{f_desc}' adicionado!")
            for k in ("f_cartao", "f_desc", "f_subtipo_santander", "f_subtipo_itau",
                      "f_cat", "f_val", "f_divide", "f_pessoa", "f_vt"):
                st.session_state.pop(k, None)
            st.rerun()

    if fixos.empty:
        st.info("Nenhum fixo cadastrado ainda.")
    else:
        if "editando_fixo_id" not in st.session_state:
            st.session_state.editando_fixo_id = None

        ativos_df = fixos[fixos["ativo"] == True]
        st.markdown(f"<small style='color:#888'>{len(ativos_df)} ativos · estimativa mensal: <b>R$ {float(ativos_df['valor_estimado'].sum()):,.2f}</b></small>", unsafe_allow_html=True)

        def _cell_fixo(row):
            ativo = bool(row.get("ativo", True))
            subtipo_val = row.get("subtipo_cartao")
            subtipo_str = str(subtipo_val) if subtipo_val and not pd.isna(subtipo_val) else None
            badge = bank_badge(str(row["cartao"]), subtipo_str)
            cat = str(row.get("categoria", ""))
            val_est = float(row.get("valor_estimado") or 0)
            pt = row.get("pessoa_thais")
            vt = row.get("valor_thais")
            pessoa_info = ""
            if pt and not pd.isna(pt):
                vt_fmt = f" · R$ {float(vt):,.2f}" if vt and not pd.isna(vt) else ""
                pessoa_info = f'<div style="font-size:10px;color:#666;margin-top:1px">👤 {pt}{vt_fmt}</div>'
            status_dot = ('<span style="color:#4ade80;font-size:11px">● ativo</span>' if ativo
                          else '<span style="color:#555;font-size:11px">● pausado</span>')
            op = "1" if ativo else "0.45"
            return (
                f'<div style="display:flex;align-items:center;font-family:inherit;padding:2px 0;opacity:{op}">'
                f'<div style="flex:3;min-width:0"><span style="font-weight:500;font-size:13px">{_fmt_desc(row["descricao"])}</span>&nbsp;{badge}{pessoa_info}</div>'
                f'<div style="flex:1.3;font-size:12px;color:#aaa">{cat}</div>'
                f'<div style="flex:1.2;text-align:right;font-size:13px;font-weight:600">R$ {val_est:,.2f}</div>'
                f'<div style="flex:1.1;text-align:center">{status_dot}</div>'
                f'</div>'
            )

        hc = st.columns([6, 0.6, 0.6, 1.4])
        hc[0].html(
            '<div style="display:flex;color:#666;font-size:11px;text-transform:uppercase;'
            'letter-spacing:.6px;font-weight:500;border-bottom:2px solid rgba(255,255,255,0.12);padding:6px 0">'
            '<div style="flex:3">Descrição</div><div style="flex:1.3">Categoria</div>'
            '<div style="flex:1.2;text-align:right">Valor Est.</div>'
            '<div style="flex:1.1;text-align:center">Status</div></div>'
        )

        for _, row in fixos.iterrows():
            fid = int(row["id"])
            ativo = bool(row.get("ativo", True))
            rc = st.columns([6, 0.6, 0.6, 1.4])
            rc[0].html(_cell_fixo(row))
            if rc[1].button("✏️", key=f"efx_{fid}", help="Editar"):
                st.session_state.editando_fixo_id = fid
                st.rerun()
            if rc[2].button("🗑", key=f"dfx_{fid}", help="Excluir"):
                delete_fixo(fid)
                st.rerun()
            if rc[3].button("⏸" if ativo else "▶", key=f"tfx_{fid}", use_container_width=True, help="Pausar" if ativo else "Ativar"):
                update_fixo(fid, ativo=not ativo)
                st.rerun()

        # ── Formulário de edição de fixo ──────────────────────────────────────
        if st.session_state.editando_fixo_id:
            ed = fixos[fixos["id"] == st.session_state.editando_fixo_id]
            if not ed.empty:
                ed = ed.iloc[0]
                st.divider()
                st.markdown(f"##### ✏️ Editando fixo: *{ed['descricao']}*")
                ge1, ge2 = st.columns(2)
                e_cartao = ge1.selectbox("Cartão", CARTOES,
                    index=CARTOES.index(ed["cartao"]) if ed["cartao"] in CARTOES else 0, key="efx_cartao")
                e_desc = ge2.text_input("Descrição", value=str(ed["descricao"]), key="efx_desc")
                e_subtipo = None
                _sub_atual = ed.get("subtipo_cartao")
                _sub_atual = str(_sub_atual) if _sub_atual and not pd.isna(_sub_atual) else None
                if e_cartao == "Santander":
                    _idx = SUBTIPOS_SANTANDER.index(_sub_atual) if _sub_atual in SUBTIPOS_SANTANDER else 0
                    e_subtipo = st.radio("Tipo Santander", SUBTIPOS_SANTANDER, index=_idx,
                        horizontal=True, key="efx_subtipo_santander")
                elif e_cartao == "Itaú":
                    _idx = SUBTIPOS_ITAU.index(_sub_atual) if _sub_atual in SUBTIPOS_ITAU else 0
                    e_subtipo = st.radio("Bandeira Itaú", SUBTIPOS_ITAU, index=_idx,
                        horizontal=True, key="efx_subtipo_itau")
                ge3, ge4 = st.columns(2)
                e_cat = ge3.selectbox("Categoria", CATEGORIAS,
                    index=CATEGORIAS.index(ed["categoria"]) if ed["categoria"] in CATEGORIAS else 0, key="efx_cat")
                e_val = ge4.number_input("Valor estimado (R$)", min_value=0.0, step=0.01, format="%.2f",
                    value=float(ed.get("valor_estimado") or 0), key="efx_val")
                _pt = ed.get("pessoa_thais")
                _tem_p = bool(_pt and not pd.isna(_pt))
                e_div = st.checkbox("Dividir?", value=_tem_p, key="efx_div")
                e_pessoa, e_vt = None, None
                if e_div:
                    gp1, gp2 = st.columns(2)
                    e_pessoa = gp1.text_input("Pessoa", value=str(_pt) if _tem_p else "Thais", key="efx_pessoa")
                    e_vt = gp2.number_input("Valor (R$)", min_value=0.0, step=0.01, format="%.2f",
                        value=float(ed.get("valor_thais") or 0) if not pd.isna(ed.get("valor_thais")) else 0.0, key="efx_vt")
                gb1, gb2 = st.columns(2)
                if gb1.button("💾 Salvar fixo", use_container_width=True, key="efx_salvar"):
                    update_fixo(st.session_state.editando_fixo_id,
                        cartao=e_cartao, subtipo_cartao=e_subtipo, descricao=e_desc,
                        categoria=e_cat, valor_estimado=e_val,
                        pessoa_thais=e_pessoa if e_div else None,
                        valor_thais=e_vt if e_div else None)
                    st.session_state.editando_fixo_id = None
                    st.rerun()
                if gb2.button("✕ Cancelar", use_container_width=True, key="efx_cancelar"):
                    st.session_state.editando_fixo_id = None
                    st.rerun()
            else:
                st.session_state.editando_fixo_id = None


# ══════════════════════════════════════════════════════════════════════════════
# IMPORTAR
# ══════════════════════════════════════════════════════════════════════════════
elif pagina == "Importar":
    st.title("Importar via print de fatura")
    st.caption("Envie um print da sua fatura. A IA identifica o banco, extrai os lançamentos com valores e parcelas, e você revisa antes de salvar.")

    import os, base64, json as _json

    # Chave da API — prioridade: variável de ambiente > session_state
    _SECRETS_FILE = Path(__file__).parent / ".streamlit" / "secrets.toml"

    def _load_saved_key() -> str:
        if _SECRETS_FILE.exists():
            for line in _SECRETS_FILE.read_text(encoding="utf-8").splitlines():
                if line.startswith("ANTHROPIC_API_KEY"):
                    return line.split("=", 1)[1].strip().strip('"').strip("'")
        return ""

    def _outras_linhas() -> list:
        """Linhas do secrets.toml que NÃO são ANTHROPIC_API_KEY (ex.: APP_PASSWORD)."""
        if not _SECRETS_FILE.exists():
            return []
        return [ln for ln in _SECRETS_FILE.read_text(encoding="utf-8").splitlines()
                if ln.strip() and not ln.lstrip().startswith("ANTHROPIC_API_KEY")]

    def _save_key(key: str):
        _SECRETS_FILE.parent.mkdir(parents=True, exist_ok=True)
        linhas = _outras_linhas() + [f'ANTHROPIC_API_KEY = "{key}"']
        _SECRETS_FILE.write_text("\n".join(linhas) + "\n", encoding="utf-8")

    def _delete_key():
        # Remove apenas a chave da API, preservando o resto (APP_PASSWORD etc.)
        linhas = _outras_linhas()
        if linhas:
            _SECRETS_FILE.write_text("\n".join(linhas) + "\n", encoding="utf-8")
        elif _SECRETS_FILE.exists():
            _SECRETS_FILE.unlink()

    # Prioridade: variável de ambiente > arquivo salvo > session_state
    # Sempre tenta carregar do arquivo (não bloqueia por session_state vazio)
    _persisted = os.environ.get("ANTHROPIC_API_KEY", "") or _load_saved_key()
    if _persisted:
        st.session_state["_imp_api_key"] = _persisted

    _trocar = st.session_state.get("_imp_trocar", False)
    api_key = st.session_state.get("_imp_api_key", "")

    if not api_key or _trocar:
        st.session_state["_imp_trocar"] = False
        _typed = st.text_input("Chave da API Anthropic (ANTHROPIC_API_KEY)", type="password",
                               help="Cole sua chave e pressione Enter. Fica salva localmente para próximas sessões.")
        if _typed:
            _save_key(_typed)
            st.session_state["_imp_api_key"] = _typed
            api_key = _typed
            st.success("✅ Chave salva!")
    else:
        st.success("✅ Chave da API configurada e salva localmente.")
        if st.button("🔑 Trocar chave", key="trocar_chave"):
            _delete_key()
            st.session_state["_imp_api_key"] = ""
            st.session_state["_imp_trocar"] = True
            st.rerun()

    uploaded_files = st.file_uploader(
        "Prints da fatura (até 5 imagens)",
        type=["png", "jpg", "jpeg", "webp", "heic", "heif"],
        accept_multiple_files=True,
        label_visibility="collapsed",
    )
    arquivos = list(uploaded_files) if uploaded_files else []
    if len(arquivos) > 5:
        st.warning("Máximo de 5 imagens por vez. Apenas as 5 primeiras serão analisadas.")
        arquivos = arquivos[:5]

    col_up1, col_up2 = st.columns(2)
    mes_imp = col_up1.selectbox("Mês de destino", meses_disponiveis,
                                 index=meses_disponiveis.index(mes) if mes in meses_disponiveis else 0,
                                 format_func=lambda x: MES_LABELS.get(x, x))
    col_up2.selectbox("Cartão (será substituído pelo detectado)", CARTOES, key="imp_cartao")

    def _to_jpeg_if_needed(nome, raw):
        """Converte HEIC/HEIF (ou formatos não suportados) para JPEG via PIL."""
        ext = (nome.rsplit(".", 1)[-1].lower() if "." in nome else "")
        if ext in ("png", "jpg", "jpeg", "webp"):
            return nome, raw, ext
        try:
            import io as _io
            from PIL import Image
            try:
                import pillow_heif
                pillow_heif.register_heif_opener()
            except Exception:
                pass
            img = Image.open(_io.BytesIO(raw)).convert("RGB")
            buf = _io.BytesIO()
            img.save(buf, format="JPEG", quality=90)
            return (nome.rsplit(".", 1)[0] + ".jpg"), buf.getvalue(), "jpeg"
        except Exception:
            return nome, raw, ext or "jpeg"

    if arquivos:
        _imgs = []
        for f in arquivos:
            nome = getattr(f, "name", "foto.jpg")
            nm, by, ex = _to_jpeg_if_needed(nome, f.read())
            _imgs.append({"name": nm, "bytes": by, "ext": ex})
        cols_prev = st.columns(len(_imgs))
        for col, img in zip(cols_prev, _imgs):
            col.image(img["bytes"], caption=img["name"], width=160)
    else:
        _imgs = []

    st.divider()
    _btn_disabled = not api_key or not _imgs
    _btn_tip = ("Configure a chave da API primeiro." if not api_key
                else "Envie ao menos uma imagem." if not _imgs else "")
    if _btn_tip and _btn_disabled:
        st.caption(f"⚠️ {_btn_tip}")

    if st.button("🔍 Analisar com IA", use_container_width=True, type="primary",
                 disabled=_btn_disabled):
        if True:
            import anthropic as _anthropic
            client = _anthropic.Anthropic(api_key=api_key)

            _prompt = """Analise este print de fatura de cartão de crédito e extraia TODOS os lançamentos visíveis.

Retorne APENAS um JSON válido com esta estrutura (sem texto extra, sem markdown):
{
  "banco": "Santander|Itaú|C6",
  "subtipo": "Virtual|Físico|null",
  "lancamentos": [
    {
      "descricao": "nome do lançamento",
      "valor": 0.00,
      "parcela_atual": 1,
      "total_parcelas": 1,
      "categoria_sugerida": "Essencial|Não essencial|Estudos|Lazer|Viagem|Reforma|Negócios|Metinha|Livre"
    }
  ]
}

Regras:
- valor deve ser positivo (número float, sem R$)
- Se não houver parcelamento, parcela_atual=1 e total_parcelas=1
- Se aparecer "2/5" ou "Parc 2 de 5", parcela_atual=2 e total_parcelas=5
- total_parcelas representa as parcelas RESTANTES incluindo a atual
- banco: identifique pelo visual, cores ou nome visível (Santander=vermelho, Itaú=laranja, C6=preto)
- subtipo: "Físico" se for cartão físico Santander, "Virtual" para virtual/padrão, null para outros bancos
- categoria_sugerida: escolha a mais adequada pelo nome do estabelecimento"""

            todos_lancamentos = []
            banco_det = None
            subtipo_det = None
            erros = []

            prog = st.progress(0, text="Analisando imagens...")
            for idx, f in enumerate(_imgs):
                prog.progress(idx / len(_imgs), text=f"Analisando {f['name']}…")
                try:
                    img_bytes = f["bytes"]
                    ext = f["ext"]
                    mime = {"png": "image/png", "jpg": "image/jpeg",
                            "jpeg": "image/jpeg", "webp": "image/webp"}.get(ext, "image/png")
                    img_b64 = base64.standard_b64encode(img_bytes).decode()
                    resp = client.messages.create(
                        model="claude-sonnet-4-6",
                        max_tokens=2000,
                        messages=[{"role": "user", "content": [
                            {"type": "image", "source": {"type": "base64", "media_type": mime, "data": img_b64}},
                            {"type": "text", "text": _prompt},
                        ]}],
                    )
                    raw = resp.content[0].text.strip()
                    if raw.startswith("```"):
                        raw = raw.split("```")[1]
                        if raw.startswith("json"):
                            raw = raw[4:]
                    dados = _json.loads(raw)
                    if not banco_det:
                        banco_det = dados.get("banco")
                        subtipo_det = dados.get("subtipo") or None
                    todos_lancamentos.extend(dados.get("lancamentos", []))
                except Exception as e:
                    erros.append(f"{f.name}: {e}")

            prog.progress(1.0, text="Concluído!")

            if erros:
                for err in erros:
                    st.error(f"Erro em {err}")
            if todos_lancamentos:
                st.session_state["imp_dados"] = {"banco": banco_det, "subtipo": subtipo_det,
                                                   "lancamentos": todos_lancamentos}
                st.session_state["imp_mes"] = mes_imp
                st.rerun()

    # ── Tabela de revisão ─────────────────────────────────────────────────────
    if "imp_dados" in st.session_state and st.session_state["imp_dados"]:
        dados = st.session_state["imp_dados"]
        lancamentos_ia = dados.get("lancamentos", [])
        banco_det = dados.get("banco") or st.session_state.get("imp_cartao", "Santander")
        subtipo_det = dados.get("subtipo") or None

        st.divider()

        # Reseta a seleção de banco/bandeira só quando chega um novo lote da IA
        if st.session_state.get("imp_banco_src") != id(lancamentos_ia):
            st.session_state["imp_banco_sel"] = banco_det if banco_det in CARTOES else "Santander"
            if banco_det == "Santander":
                st.session_state["imp_subtipo_santander"] = subtipo_det or "Virtual"
            elif banco_det == "Itaú":
                st.session_state["imp_subtipo_itau"] = subtipo_det or "Visa"
            st.session_state["imp_banco_src"] = id(lancamentos_ia)

        st.markdown("**Banco detectado** — corrija aqui se a IA identificou errado:")
        cb1, cb2 = st.columns(2)
        banco_sel = cb1.selectbox("Cartão", CARTOES, key="imp_banco_sel")
        subtipo_sel = None
        if banco_sel == "Santander":
            if "imp_subtipo_santander" not in st.session_state:
                st.session_state["imp_subtipo_santander"] = "Virtual"
            subtipo_sel = cb2.radio("Tipo Santander", SUBTIPOS_SANTANDER,
                                     horizontal=True, key="imp_subtipo_santander")
        elif banco_sel == "Itaú":
            if "imp_subtipo_itau" not in st.session_state:
                st.session_state["imp_subtipo_itau"] = "Visa"
            subtipo_sel = cb2.radio("Bandeira Itaú", SUBTIPOS_ITAU,
                                     horizontal=True, key="imp_subtipo_itau")
        badge_det = bank_badge(banco_sel, subtipo_sel)
        st.html(f'<div style="font-size:14px;margin:4px 0 8px">{badge_det}&nbsp;&nbsp;<span style="color:#888;font-size:12px">{len(lancamentos_ia)} lançamento(s) encontrado(s)</span></div>')

        pessoa_imp = st.text_input("Pessoa (quem paga a parte preenchida em \"Thais paga\")",
                                    value=st.session_state.get("imp_pessoa") or "Thais", key="imp_pessoa",
                                    help="Padrão Thais, mas pode trocar por outra pessoa (ex.: Mãe) "
                                         "— vale para todos os itens com valor preenchido na coluna abaixo.")

        # Carrega lançamentos existentes para deduplicação
        from difflib import SequenceMatcher
        lanc_existentes = get_lancamentos(st.session_state.get("imp_mes", mes_imp))

        def _dup_check(desc: str, valor: float):
            """Retorna (é_dup, motivo) comparando valor exato e similaridade de descrição."""
            if lanc_existentes.empty:
                return False, ""
            desc_norm = desc.strip().lower()
            for _, ex in lanc_existentes.iterrows():
                ex_desc = str(ex["descricao"]).strip().lower()
                ex_val = float(ex["valor"])
                valor_igual = abs(ex_val - valor) < 0.01
                similaridade = SequenceMatcher(None, desc_norm, ex_desc).ratio()
                if valor_igual and similaridade >= 0.85:
                    return True, f"idêntico a '{ex['descricao']}' (R$ {ex_val:,.2f})"
                if valor_igual:
                    return True, f"mesmo valor que '{ex['descricao']}' (R$ {ex_val:,.2f})"
                if similaridade >= 0.85:
                    return True, f"descrição similar a '{ex['descricao']}' (R$ {ex_val:,.2f})"
            return False, ""

        # Inicializa estado editável
        if "imp_rows" not in st.session_state or st.session_state.get("imp_rows_src") != id(lancamentos_ia):
            rows_init = []
            for l in lancamentos_ia:
                is_dup, motivo = _dup_check(l["descricao"], float(l["valor"]))
                rows_init.append({**l, "_ativo": not is_dup, "_dup": is_dup, "_dup_motivo": motivo,
                                   "_fixo": False, "_valor_thais": 0.0})
            st.session_state["imp_rows"] = rows_init
            st.session_state["imp_rows_src"] = id(lancamentos_ia)

        rows = st.session_state["imp_rows"]

        # Cabeçalho
        col_widths = [0.4, 2.1, 1.2, 0.6, 0.8, 1.0, 0.6, 0.6, 1.1]
        _nome_pessoa_imp = pessoa_imp.strip() or "Thais"
        col_labels = ["✓", "Descrição", "Categoria", "Fixo", "Valor", f"{_nome_pessoa_imp} paga (R$)", "Parc.", "Total", "Tipo"]
        hc = st.columns(col_widths)
        for col, label in zip(hc, col_labels):
            col.markdown(f"<small style='color:#666;text-transform:uppercase;letter-spacing:.5px;font-size:11px'>{label}</small>", unsafe_allow_html=True)

        st.divider()
        for i, row in enumerate(rows):
            is_dup = row.get("_dup", False)
            rc = st.columns(col_widths)
            row["_ativo"] = rc[0].checkbox("", value=row["_ativo"], key=f"imp_ck_{i}", label_visibility="collapsed")
            dup_tip = f"⚠️ Possível duplicata: {row.get('_dup_motivo','já existe')}" if is_dup else ""
            row["descricao"] = rc[1].text_input("", value=row["descricao"], key=f"imp_desc_{i}",
                                                  label_visibility="collapsed", help=dup_tip)
            cats_idx = CATEGORIAS.index(row["categoria_sugerida"]) if row.get("categoria_sugerida") in CATEGORIAS else 0
            row["categoria_sugerida"] = rc[2].selectbox("", CATEGORIAS, index=cats_idx,
                                                          key=f"imp_cat_{i}", label_visibility="collapsed")
            row["_fixo"] = rc[3].checkbox("Fixo", value=row.get("_fixo", False), key=f"imp_fixo_{i}",
                                           label_visibility="collapsed",
                                           help="Marcar se é um gasto fixo novo (recorrente todo mês). "
                                                "Vai também para a tela Fixos depois de importar.")
            row["valor"] = rc[4].number_input("", value=float(row["valor"]), min_value=0.0,
                                               step=0.01, format="%.2f", key=f"imp_val_{i}",
                                               label_visibility="collapsed")
            row["_valor_thais"] = rc[5].number_input("", value=float(row.get("_valor_thais") or 0.0),
                                                       min_value=0.0, step=0.01, format="%.2f",
                                                       key=f"imp_valthais_{i}", label_visibility="collapsed",
                                                       help="Quanto a Thais (ou outra pessoa) vai pagar deste item, se aplicável.")
            row["parcela_atual"] = rc[6].number_input("", value=int(row.get("parcela_atual", 1)),
                                                        min_value=1, step=1, key=f"imp_parc_{i}",
                                                        label_visibility="collapsed")
            row["total_parcelas"] = rc[7].number_input("", value=int(row.get("total_parcelas", 1)),
                                                         min_value=1, step=1, key=f"imp_tot_{i}",
                                                         label_visibility="collapsed")
            tot = int(row["total_parcelas"])
            parc = int(row["parcela_atual"])
            restantes = tot - parc + 1
            tipo = "FIXO" if (row["_fixo"] or tot > 90) else ("ULTIMA" if restantes == 1 else ("única" if tot == 1 else "parcelado"))
            rc[8].markdown(f"`{tipo}`" + (" ⚠️ dup" if is_dup else ""), unsafe_allow_html=True)

        n_ativos = sum(1 for r in rows if r["_ativo"])
        n_dup = sum(1 for r in rows if r["_dup"] and r["_ativo"])

        st.divider()
        si1, si2 = st.columns([3, 1])
        si1.markdown(f"**{n_ativos}** selecionado(s)" + (f" · ⚠️ {n_dup} possível(is) duplicata(s)" if n_dup else ""))

        if si2.button("💾 Importar selecionados", use_container_width=True, type="primary", disabled=n_ativos == 0):
            mes_dest = st.session_state.get("imp_mes", mes_imp)
            banco_final = banco_sel
            subtipo_final = subtipo_sel
            importados = 0
            fixos_criados = 0
            for row in rows:
                if not row["_ativo"]:
                    continue
                tot = int(row["total_parcelas"])
                parc = int(row["parcela_atual"])
                restantes = tot - parc + 1
                is_fixo = bool(row.get("_fixo"))
                tipo = "FIXO" if (is_fixo or tot > 90) else ("ULTIMA" if restantes == 1 else ("única" if tot == 1 else "parcelado"))

                val_thais = float(row.get("_valor_thais") or 0.0)
                pessoa_thais = pessoa_imp if (val_thais > 0 and pessoa_imp.strip()) else None
                valor_thais = val_thais if val_thais > 0 else None

                if tipo == "parcelado" and restantes > 1:
                    # Cria grupo de parcelamento a partir do mês destino
                    criar_grupo_parcelamento(
                        descricao=row["descricao"],
                        cartao=banco_final,
                        subtipo_cartao=subtipo_final,
                        categoria=row["categoria_sugerida"],
                        valor_parcela=float(row["valor"]),
                        total_parcelas=restantes,
                        mes_inicio=mes_dest,
                        pessoa_thais=pessoa_thais,
                        valor_thais=valor_thais,
                    )
                else:
                    add_lancamento(
                        mes_ano=mes_dest,
                        cartao=banco_final,
                        dono="Kelvin",
                        valor=float(row["valor"]),
                        descricao=row["descricao"],
                        categoria=row["categoria_sugerida"],
                        tipo_parcela=tipo,
                        parcela_atual=parc,
                        total_parcelas=restantes,
                        subtipo_cartao=subtipo_final,
                        pessoa_thais=pessoa_thais,
                        valor_thais=valor_thais,
                    )
                importados += 1

                if is_fixo:
                    df_fixos = load_sheet("fixos")
                    novo_id = int(df_fixos["id"].max() + 1) if not df_fixos.empty else 1
                    novo_fixo = {
                        "id": novo_id, "cartao": banco_final, "subtipo_cartao": subtipo_final,
                        "descricao": row["descricao"], "categoria": row["categoria_sugerida"],
                        "valor_estimado": float(row["valor"]),
                        "pessoa_thais": pessoa_thais, "valor_thais": valor_thais,
                        "ativo": True,
                    }
                    df_fixos = pd.concat([df_fixos, pd.DataFrame([novo_fixo])], ignore_index=True)
                    save_sheet("fixos", df_fixos)
                    fixos_criados += 1

            del st.session_state["imp_dados"]
            del st.session_state["imp_rows"]
            if fixos_criados:
                st.session_state["_nav_redirect"] = "Fixos"
                st.success(f"✅ {importados} lançamento(s) importado(s) — {fixos_criados} novo(s) fixo(s) "
                           f"cadastrado(s). Indo para a tela Fixos…")
            else:
                st.success(f"✅ {importados} lançamento(s) importado(s) para {MES_LABELS.get(mes_dest, mes_dest)}!")
            st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURAÇÕES
# ══════════════════════════════════════════════════════════════════════════════
elif pagina == "Configurações":
    st.title("Configurações")
    cfg = get_config()

    # ── Nomes e divisão ───────────────────────────────────────────────────────
    # Conta da Mãe não tem "parceiro(a)" nem divisão de despesas compartilhadas —
    # é só o nome dela mesma, usado no rótulo do salário no Dashboard.
    if get_usuario_atual() == "mae":
        st.markdown("#### Perfil")
        nome_k = st.text_input("Seu nome", value=cfg.get("nome_kelvin", "Rita"))
        if st.button("💾", help="Salvar nome"):
            set_config("nome_kelvin", nome_k)
            st.success("Nome salvo!")
            st.rerun()
    else:
        st.markdown("#### Perfil")
        cn1, cn2 = st.columns(2)
        nome_k = cn1.text_input("Seu nome", value=cfg.get("nome_kelvin", "Kelvin"))
        nome_t = cn2.text_input("Nome do(a) parceiro(a)", value=cfg.get("nome_thais", "Thais"))

        st.markdown("#### Divisão de despesas compartilhadas")
        cd1, cd2 = st.columns(2)
        div_k = cd1.number_input(f"% de {nome_k}", min_value=0, max_value=100,
                                  value=int(cfg.get("divisao_kelvin", 80)), step=1)
        div_t = cd2.number_input(f"% de {nome_t}", min_value=0, max_value=100,
                                  value=int(cfg.get("divisao_thais", 20)), step=1)

        if div_k + div_t != 100:
            st.warning(f"A soma deve ser 100%. Atualmente: {div_k + div_t}%")

        if st.button("💾 Salvar", disabled=(div_k + div_t != 100), help="Salvar configurações"):
            set_config("nome_kelvin", nome_k)
            set_config("nome_thais", nome_t)
            set_config("divisao_kelvin", div_k)
            set_config("divisao_thais", div_t)
            st.success("Configurações salvas!")
            st.rerun()

    # ── Criar meses futuros (previsão) ────────────────────────────────────────
    st.divider()
    st.markdown("#### Criar meses futuros")
    _meses_atuais = get_meses()
    _ultimo = _meses_atuais[-1] if _meses_atuais else "2026-12"
    st.caption(f"Último mês cadastrado: **{fmt_mes(_ultimo)}**. "
               "Crie quantos meses quiser à frente para projetar suas finanças.")

    cm1, cm2 = st.columns([1, 2])
    qtd_meses = cm1.number_input("Quantos meses adicionar", min_value=1, max_value=36, value=1, step=1)
    aplicar_fixos_novo = cm2.checkbox(
        "Já lançar os gastos fixos ativos em cada novo mês", value=True,
        help="Projeta automaticamente os fixos (com valor estimado) nos meses criados.")

    # Prévia dos meses que serão criados
    _previa = [_proximo_mes(_ultimo, i + 1) for i in range(int(qtd_meses))]
    st.caption("Serão criados: " + " · ".join(fmt_mes(m) for m in _previa))

    if st.button("➕ Criar meses", use_container_width=False):
        sal_base = get_mes(_ultimo)
        sk = float(sal_base.get("salario_kelvin") or 0)
        stz = float(sal_base.get("salario_thais") or 0)
        novos_meses = [m for m in _previa if m not in _meses_atuais]
        for m in novos_meses:
            upsert_mes(m, sk, stz)            # herda salários do último mês (editável depois)

        # Projeta fixos em lote (1 escrita só) — evita lentidão/concorrência
        if aplicar_fixos_novo and novos_meses:
            fixos_ativos = get_fixos(apenas_ativos=True)
            linhas = []
            for m in novos_meses:
                for _, fx in fixos_ativos.iterrows():
                    vt = float(fx["valor_thais"]) if fx.get("valor_thais") and not pd.isna(fx["valor_thais"]) else None
                    pt = str(fx["pessoa_thais"]) if fx.get("pessoa_thais") and not pd.isna(fx["pessoa_thais"]) else None
                    linhas.append({
                        "mes_ano": m, "cartao": str(fx["cartao"]), "dono": "Kelvin",
                        "valor": float(fx.get("valor_estimado") or 0),
                        "descricao": str(fx["descricao"]), "categoria": str(fx["categoria"]),
                        "valor_thais": vt, "pessoa_thais": pt, "tipo_parcela": "FIXO",
                    })
            add_lancamentos_bulk(linhas)

        st.success(f"{len(novos_meses)} mês(es) criado(s)! Salários herdados de {fmt_mes(_ultimo)} "
                   f"(ajuste em cada mês na Home).")
        st.rerun()

    # ── Exportação completa ───────────────────────────────────────────────────
    st.divider()
    st.markdown("#### Exportar todos os meses")
    st.caption("Gera um arquivo Excel com uma aba por mês contendo todos os lançamentos.")

    if st.button("📥 Gerar exportação completa", use_container_width=False):
        import io
        from openpyxl import Workbook as _WB
        from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align
        wb_full = _WB()
        wb_full.remove(wb_full.active)
        todos_meses = get_meses()
        for m in todos_meses:
            df_m = get_lancamentos(m)
            label_m = MES_LABELS.get(m, m).replace("/", "-")
            ws = wb_full.create_sheet(label_m)
            cols_exp = ["descricao", "categoria", "cartao", "subtipo_cartao",
                        "valor", "tipo_parcela", "total_parcelas", "pessoa_thais", "valor_thais"]
            headers = ["Descrição", "Categoria", "Cartão", "Subtipo", "Valor",
                       "Tipo", "Faltam", "Pessoa", "Valor Pessoa"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = _Font(bold=True, color="FFFFFF")
                cell.fill = _Fill("solid", start_color="1A1A2E")
                cell.alignment = _Align(horizontal="center")
            if not df_m.empty:
                for _, row in df_m.iterrows():
                    ws.append([row.get(c, "") for c in cols_exp])
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 18
        buf = io.BytesIO()
        wb_full.save(buf)
        buf.seek(0)
        st.download_button("⬇️ Baixar Excel completo", data=buf,
            file_name="controle_financeiro_completo.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False)

    # ── Backup ────────────────────────────────────────────────────────────────
    st.divider()
    st.markdown("#### Backup do banco de dados")

    from modules.db import DB_PATH, USE_POSTGRES
    from datetime import datetime as _dt

    if USE_POSTGRES:
        st.info("🗄️ Backend: **Postgres (Supabase)**. Os backups são gerenciados "
                "automaticamente pela plataforma. Use a exportação acima para uma cópia em Excel.")

    backups = listar_backups()
    bc1, bc2 = st.columns([3, 1])
    bc1.caption("Local: Supabase Postgres" if USE_POSTGRES else f"Local: `{DB_PATH}`")
    if not USE_POSTGRES:
        if bc2.button("💾 Fazer backup agora", use_container_width=True):
            caminho = fazer_backup()
            st.success(f"Backup salvo: `{Path(caminho).name}`")
            st.rerun()

    if backups:
        rows_bk = ""
        for i, b in enumerate(backups):
            ts_str = b.stem.replace("financeiro_", "")
            try:
                ts_fmt = _dt.strptime(ts_str, "%Y%m%d_%H%M%S").strftime("%d/%m/%Y %H:%M:%S")
            except Exception:
                ts_fmt = ts_str
            size_kb = b.stat().st_size / 1024
            bg = "rgba(255,255,255,0.02)" if i % 2 == 0 else "transparent"
            tag = ' <span style="background:#1a3a2a;color:#4ade80;padding:1px 6px;border-radius:3px;font-size:10px">mais recente</span>' if i == 0 else ""
            rows_bk += f'<tr style="background:{bg};border-bottom:1px solid rgba(255,255,255,0.06)"><td style="padding:8px 14px;font-size:13px">{ts_fmt}{tag}</td><td style="padding:8px 14px;text-align:right;font-size:12px;color:#888">{size_kb:.1f} KB</td></tr>'
        st.html(f"""<table style="width:100%;border-collapse:collapse;font-family:inherit">
          <thead><tr style="border-bottom:2px solid rgba(255,255,255,0.12)">
            <th style="padding:8px 14px;text-align:left;font-size:11px;color:#666;text-transform:uppercase;letter-spacing:.6px">Data</th>
            <th style="padding:8px 14px;text-align:right;font-size:11px;color:#666;text-transform:uppercase;letter-spacing:.6px">Tamanho</th>
          </tr></thead><tbody>{rows_bk}</tbody></table>""")
        st.caption(f"Últimos {len(backups)} backup(s) mantidos automaticamente (máx. 7).")
    elif not USE_POSTGRES:
        st.caption("Nenhum backup encontrado.")
