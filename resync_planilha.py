"""
Re-sincroniza o Supabase com a planilha CONTAS.xlsx (meses Jul-26 a Dez-26).

- Faz BACKUP local dos lançamentos e grupos atuais antes de mexer.
- Reimporta os lançamentos dos 6 meses a partir da planilha (fonte da verdade).
- PRESERVA os parcelamentos criados pelo app (linhas com id_grupo) e seus grupos.
- Atualiza salários (meses) e o painel-resumo (água, investimentos, lembretes).

Uso:
    set DATABASE_URL=postgresql://...
    venv\\Scripts\\python.exe resync_planilha.py
"""
import os
import sys
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
import modules.db as d

ORIGEM = Path(r"C:\Users\kelvin.araujo\Downloads\CONTAS.xlsx")
MES_MAP = {
    "Julho-26": "2026-07", "Agosto-26": "2026-08", "Setembro-26": "2026-09",
    "Outubro-26": "2026-10", "Novembro-26": "2026-11", "Dezembro-26": "2026-12",
}
CARTOES_VALIDOS = {"Itaú", "Santander", "C6"}
NORMALIZA_CATEGORIA = {
    "viagen": "Viagem", "viagem": "Viagem", "essencial": "Essencial",
    "não essencial": "Não essencial", "estudos": "Estudos", "lazer": "Lazer",
    "reforma": "Reforma", "negócios": "Negócios", "metinha": "Metinha", "livre": "Livre",
}
COLS = d._SCHEMA["lancamentos"]


def _cat(v):
    if not v or not isinstance(v, str):
        return "Não essencial"
    return NORMALIZA_CATEGORIA.get(v.strip().lower(), v.strip())


def _tipo(v):
    if isinstance(v, str):
        u = v.strip().upper()
        if u in ("FIXO", "ULTIMA"):
            return u, None, None
        return "única", None, None
    if isinstance(v, (int, float)):
        return "parcelado", None, int(v)
    return "única", None, None


def main():
    if not d.USE_POSTGRES:
        print("ERRO: defina DATABASE_URL (Supabase) antes de rodar.")
        sys.exit(1)
    if not ORIGEM.exists():
        print(f"ERRO: planilha não encontrada: {ORIGEM}")
        sys.exit(1)

    # 1) Backup
    bkp_dir = Path(__file__).parent / "data" / "backups"
    bkp_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    old_lanc = d.load_sheet("lancamentos")
    old_grp = d.load_sheet("grupos_parcelamento")
    bkp = bkp_dir / f"supabase_pre_resync_{ts}.xlsx"
    with pd.ExcelWriter(bkp) as xw:
        old_lanc.to_excel(xw, sheet_name="lancamentos", index=False)
        old_grp.to_excel(xw, sheet_name="grupos_parcelamento", index=False)
    print(f"Backup salvo em: {bkp}")

    # 2) Preserva parcelamentos do app (id_grupo preenchido)
    if not old_lanc.empty and "id_grupo" in old_lanc.columns:
        preservar = old_lanc[old_lanc["id_grupo"].notna()].copy()
    else:
        preservar = pd.DataFrame(columns=COLS)
    print(f"Parcelamentos do app preservados: {len(preservar)} linha(s)")

    # 3) Reimporta os 6 meses da planilha
    wb = load_workbook(ORIGEM, data_only=True)
    novos = []
    for aba, ma in MES_MAP.items():
        if aba not in wb.sheetnames:
            continue
        ws = wb[aba]
        d.upsert_mes(ma, float(ws["K4"].value or 0), float(ws["N4"].value or 0))
        # painel
        d.set_painel(ma,
                     agua_boleto=float(ws["L10"].value or 0),
                     youtube_lembrete=str(ws["K14"].value or ""),
                     spotify_lembrete=str(ws["K15"].value or ""),
                     cdb_reserva=float(ws["L18"].value or 0),
                     previdencia=float(ws["L19"].value or 0))
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 9:
                continue
            cartao, dono, valor, desc, cat, vt, pt, tipo_raw = (
                row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8])
            if cartao not in CARTOES_VALIDOS:
                continue
            if not isinstance(valor, (int, float)) or valor == 0 or not desc:
                continue
            tp, pa, tot = _tipo(tipo_raw)
            novos.append({
                "mes_ano": ma, "cartao": str(cartao),
                "dono": str(dono) if dono else "Kelvin", "valor": float(valor),
                "descricao": str(desc).strip(), "categoria": _cat(cat),
                "valor_thais": float(vt) if isinstance(vt, (int, float)) else None,
                "pessoa_thais": str(pt).strip() if isinstance(pt, str) and pt.strip() not in ("", "None") else None,
                "tipo_parcela": tp, "parcela_atual": pa, "total_parcelas": tot,
                "id_grupo": None, "subtipo_cartao": None,
                "data_lancamento": None, "conferido": False,
            })
        print(f"  {ma}: {sum(1 for x in novos if x['mes_ano']==ma)} linhas da planilha")

    df_novos = pd.DataFrame(novos)

    # 4) Junta planilha + parcelamentos preservados, re-IDs
    if not preservar.empty:
        for c in COLS:
            if c not in preservar.columns:
                preservar[c] = None
        preservar = preservar[COLS]
        df_final = pd.concat([df_novos, preservar.drop(columns=["id"])], ignore_index=True)
    else:
        df_final = df_novos
    df_final.insert(0, "id", range(1, len(df_final) + 1))
    df_final = df_final[COLS]
    if "conferido" in df_final.columns:
        df_final["conferido"] = df_final["conferido"].fillna(False).astype(bool)

    d.save_sheet("lancamentos", df_final)
    print(f"\nLançamentos no Supabase: {len(df_final)} (planilha {len(df_novos)} + preservados {len(preservar)})")
    print("Re-sync concluído.")


if __name__ == "__main__":
    main()
